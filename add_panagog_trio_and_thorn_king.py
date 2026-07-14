"""
add_panagog_trio_and_thorn_king.py — Add four boss/NPC creatures to
combat tracker.xlsx as creature tabs + Battle List rows:

  - Korvath Stonehorn   (Shield of the Circle — minotaur heavy warrior, CR 9)
  - Varka Ashstride     (Knife on the Wind — orc mobile warrior, CR 8)
  - Sseskari Reed-Speaker (Voice Beneath the Stone — lizardfolk shaman, CR 9)
  - Tharok Ghostclaw    (The Thorn King, Trial I — apex predator, CR 14)

Format matches engine/excel_loader.py exactly (fixed-cell stat blocks +
TRAITS/ACTIONS/BONUS ACTIONS/REACTIONS/BLOODIED section anchors, 2-column
Name/Description rows). Helper functions lifted verbatim from add_npcs.py /
add_monster.py. Recharge-limited attacks with a to-hit/damage/save live in
the ATTACKS table (Type = "Special (Recharge X-Y)"); Multiattack, Legendary
Resistance, Legendary Actions, and innate/spellcasting summaries live in
TRAITS — matching the convention used by Adult Red Dragon / Storm Giant /
Wraith.

SAFE WRITE: copies the workbook to /tmp, modifies the copy, then replaces
the original. Existing sheets are never deleted unless a new tab collides
by name (none of these do). Battle List rows are appended with Quantity 0
(library entries — bump to >0 to drop one into a fight).

RUN from the battle tracker folder:  python3 add_panagog_trio_and_thorn_king.py
"""
import openpyxl
import shutil
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
WORKBOOK = "/sessions/keen-dazzling-allen/mnt/battle tracker/combat tracker.xlsx"
TMP = "/tmp/combat_tracker_bosses.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS (verbatim format from add_npcs.py / add_monster.py)
# ─────────────────────────────────────────────────────────────────────────────


def write_section(ws, start_row, anchor, entries, col_b_label="Description"):
    ws.cell(row=start_row, column=1, value=anchor)
    ws.cell(row=start_row + 1, column=1, value="Name")
    ws.cell(row=start_row + 1, column=2, value=col_b_label)
    r = start_row + 2
    for name, desc in entries:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=desc)
        r += 1
    return r + 1  # blank separator row


def add_creature_tab(wb, c):
    tab = c["tab"]
    if tab in wb.sheetnames:
        del wb[tab]
    ws = wb.create_sheet(tab)

    ws["A1"] = "CREATURE STAT BLOCK"
    ws["K2"] = c.get("atk_per_round", 1)

    ws["A3"] = "Name"
    ws["B3"] = tab
    ws["A4"] = "Epithet"
    ws["B4"] = c.get("epithet", "")
    ws["A5"] = "Role"
    ws["B5"] = c.get("role", "")
    ws["A6"] = "Size / Type / Alignment"
    ws["B6"] = c.get("size_type", "")
    ws["A7"] = "AC"
    ws["B7"] = c["ac"]
    ws["A8"] = "Base HP"
    ws["B8"] = c["hp"]
    ws["A9"] = "Speed"
    ws["B9"] = c.get("speed", "30 ft.")
    ws["A10"] = "Proficiency Bonus"
    ws["B10"] = c.get("prof", 2)
    ws["A11"] = "Initiative Mod"
    ws["B11"] = c.get("init_mod", 0)
    ws["A12"] = "CR"
    ws["B12"] = c.get("cr", "")

    # Ability scores
    ws["A13"] = "ABILITY SCORES"
    for i, lbl in enumerate(["STR", "DEX", "CON", "INT", "WIS", "CHA"]):
        ws.cell(row=14, column=2 + i, value=lbl)
    ws["A15"] = "Score"
    ws["A16"] = "Modifier"
    ws["A17"] = "Save Proficient"
    ws["A18"] = "Save Bonus"
    for i, (sc, mo, sp, sv) in enumerate(
        zip(c["scores"], c["mods"], c.get("save_prof", [None] * 6), c["saves"])
    ):
        ws.cell(row=15, column=2 + i, value=sc)
        ws.cell(row=16, column=2 + i, value=mo)
        if sp:
            ws.cell(row=17, column=2 + i, value="Y")
        ws.cell(row=18, column=2 + i, value=sv)

    # Defenses
    ws["A20"] = "DEFENSES & SENSES"
    ws["A21"] = "Resistances"
    ws["B21"] = c.get("resistances", "")
    ws["A22"] = "Immunities"
    ws["B22"] = c.get("immunities", "")
    ws["A23"] = "Condition Immunities"
    ws["B23"] = c.get("cond_immune", "")
    ws["A24"] = "Senses"
    ws["B24"] = c.get("senses", "")

    # Attacks
    ws["A26"] = "ATTACKS"
    for col, lbl in enumerate(
        ["Name", "Type", "To Hit", "Reach / Range", "Damage", "Save", "Effect"], start=1
    ):
        ws.cell(row=27, column=col, value=lbl)
    r = 28
    for nm, tp, hit, rng, dmg, sv, eff in c.get("attacks", []):
        ws.cell(row=r, column=1, value=nm)
        ws.cell(row=r, column=2, value=tp)
        ws.cell(row=r, column=3, value=hit)
        ws.cell(row=r, column=4, value=rng)
        ws.cell(row=r, column=5, value=dmg)
        ws.cell(row=r, column=6, value=sv)
        ws.cell(row=r, column=7, value=eff)
        r += 1
    next_row = r + 1

    for anchor, entries in [
        ("TRAITS", c.get("traits", [])),
        ("ACTIONS", c.get("actions", [])),
        ("BONUS ACTIONS", c.get("bonus_actions", [])),
        ("REACTIONS", c.get("reactions", [])),
        ("BLOODIED", c.get("bloodied", [])),
    ]:
        if entries:
            next_row = write_section(ws, next_row, anchor, entries)

    return ws


def add_battle_list_row(wb, c):
    bl = wb["Battle List"]
    insert_row = bl.max_row + 1
    for r in range(1, bl.max_row + 2):
        if bl.cell(row=r, column=1).value is None:
            insert_row = r
            break
    bl.cell(row=insert_row, column=1, value=c.get("bl_type", "NPC"))
    bl.cell(row=insert_row, column=2, value=c["tab"])
    bl.cell(row=insert_row, column=3, value=c["tab"])
    bl.cell(row=insert_row, column=4, value=c.get("bl_quantity", 0))
    bl.cell(row=insert_row, column=5, value=c.get("bl_group_name") or None)
    bl.cell(row=insert_row, column=6, value=c.get("bl_init_mode", "Individual"))
    bl.cell(row=insert_row, column=7, value=c.get("bl_hp_override") or None)
    bl.cell(row=insert_row, column=8, value=c.get("bl_notes") or None)
    return insert_row


# ─────────────────────────────────────────────────────────────────────────────
# CREATURES
# ─────────────────────────────────────────────────────────────────────────────

CREATURES = [
    # ── Korvath Stonehorn ────────────────────────────────────────────────────
    {
        "tab": "Korvath Stonehorn",
        "epithet": "Shield of the Circle",
        "role": "Minotaur Heavy Warrior — tank / front-liner",
        "size_type": "Large monstrosity, lawful neutral",
        "ac": 19,
        "hp": 168,
        "speed": "40 ft.",
        "prof": 4,
        "init_mod": 1,
        "cr": "9",
        "atk_per_round": 3,
        "scores": [22, 12, 20, 10, 14, 12],
        "mods": [6, 1, 5, 0, 2, 1],
        "save_prof": [True, False, True, False, True, False],
        "saves": [10, 1, 9, 0, 6, 1],
        "senses": "darkvision 60 ft., passive Perception 16",
        "attacks": [
            ("Warhammer", "Melee Weapon", 10, "10 ft.", "2d8+6 bludgeoning", "—", ""),
            (
                "Horn Bash",
                "Melee Weapon",
                10,
                "5 ft.",
                "2d6+6 piercing",
                "DC 18 STR",
                "On a failed save, target is pushed 10 ft. or knocked prone (Korvath's choice).",
            ),
            (
                "Driving Charge",
                "Special (Recharge 5-6)",
                "—",
                "Speed, in a line",
                "4d10+5 bludgeoning",
                "DC 18 DEX",
                "Korvath moves up to his speed in a straight line without provoking opportunity "
                "attacks, and may move through the space of one Large or smaller creature. That "
                "creature takes damage and falls prone on a failed save, or takes half damage on a success.",
            ),
        ],
        "traits": [
            ("Multiattack", "Korvath makes two Warhammer attacks and one Horn Bash attack."),
            ("Indomitable (1/Day)", "Korvath rerolls a failed saving throw."),
            (
                "Guardian's Reach",
                "A creature provokes an opportunity attack from Korvath when it moves from a space "
                "within 10 feet of him to another space within 10 feet of him.",
            ),
            (
                "Stand Over the Fallen",
                "Allies within 5 feet of Korvath have half cover. A prone or unconscious ally in that "
                "area has three-quarters cover.",
            ),
            (
                "Relentless (1/Day)",
                "If reduced to 0 hit points but not killed outright, Korvath drops to 1 hit point instead.",
            ),
            ("Languages", "Common, Giant, Panagog"),
            (
                "Tactics",
                "Takes the center and physically blocks access to Sseskari. Uses Horn Bash to "
                "separate a melee threat from a mage or artificer, then holds the gap. Spends actions "
                "shoving, blocking, or protecting rather than maximizing damage. If an ally falls, moves "
                "over them and demands space for surrender or healing. Never attacks a yielded foe; only "
                "attacks an unconscious foe if it keeps returning as an active threat. Demeanor: quiet, "
                "formal, steady — respects anyone who holds position to protect another. Yields if both "
                "allies are removed and he is below 40 HP, or if continuing would abandon a dying "
                "companion; accepts surrender immediately.",
            ),
        ],
        "reactions": [
            (
                "Interpose Shield",
                "When a creature Korvath can see attacks an ally within 5 feet of him, he imposes "
                "disadvantage on the attack roll.",
            ),
        ],
        "bl_type": "NPC",
        "bl_quantity": 0,
        "bl_notes": "Panagog Trial boss — Shield of the Circle. Tanks center; PRIORITY: protects Sseskari.",
    },
    # ── Varka Ashstride ──────────────────────────────────────────────────────
    {
        "tab": "Varka Ashstride",
        "epithet": "Knife on the Wind",
        "role": "Orc Mobile Warrior — skirmisher",
        "size_type": "Medium humanoid (orc), chaotic neutral",
        "ac": 18,
        "hp": 132,
        "speed": "45 ft.",
        "prof": 4,
        "init_mod": 6,
        "cr": "8",
        "atk_per_round": 3,
        "scores": [16, 22, 16, 11, 14, 14],
        "mods": [3, 6, 3, 0, 2, 2],
        "save_prof": [False, True, True, False, True, False],
        "saves": [3, 10, 7, 0, 6, 2],
        "senses": "passive Perception 16",
        "attacks": [
            (
                "Ashfang",
                "Melee Weapon",
                10,
                "5 ft.",
                "1d8+6 slashing",
                "—",
                "Plus 7 (2d6) damage if Varka moved at least 15 feet before the attack.",
            ),
            ("Shortbow", "Ranged Weapon", 10, "80/320 ft.", "1d6+6 piercing", "—", ""),
            (
                "Hamstring Cut",
                "Special (Recharge 5-6)",
                10,
                "5 ft.",
                "2d8+6 slashing",
                "DC 17 CON",
                "On a failed save, target's speed is reduced to 0 until the end of its next turn.",
            ),
        ],
        "traits": [
            ("Multiattack", "Varka makes three attacks with Ashfang or her shortbow."),
            (
                "Evasion",
                "If Varka succeeds on a Dexterity save for half damage, she instead takes no damage; "
                "on a failure, she takes half damage.",
            ),
            (
                "Mobile Assault",
                "Varka does not provoke opportunity attacks from a creature she has attacked during "
                "the same turn.",
            ),
            (
                "Skirmisher",
                "When an enemy ends its turn within 5 feet of Varka, she may move up to 15 feet as a "
                "reaction without provoking opportunity attacks.",
            ),
            (
                "Uncanny Dodge",
                "Varka halves the damage from one attack that hits her, provided she can see the attacker.",
            ),
            (
                "Whirl Through the Line (1/Day)",
                "Varka moves up to her speed without provoking opportunity attacks and may make one "
                "Ashfang attack against up to three different creatures she moves adjacent to during "
                "this movement.",
            ),
            ("Languages", "Common, Orc, Panagog"),
            (
                "Tactics",
                "Flanks widely and pressures the mage first, unless the artificer is maintaining a "
                "dangerous concentration effect. Uses movement to force the party to split attention; "
                "doesn't stay beside the fighter unless trapped. Threatens fallen opponents to force hard "
                "choices, but prefers demanding surrender over striking them. Switches targets often to "
                "test whether the party protects one another; if Korvath pins a target, she attacks "
                "someone else instead of dogpiling (unless the party is dominating). Demeanor: "
                "sharp-tongued, competitive, delighted by capable opponents — her taunts challenge "
                "courage, not identity. Rarely yields alone; obeys a team surrender from Sseskari or "
                "Korvath. If last standing below 25 HP, offers the party one final chance to accept her "
                "surrender before a reckless final attack.",
            ),
        ],
        "bonus_actions": [
            ("Cunning Action", "Varka takes the Dash, Disengage, or Hide action."),
        ],
        "bl_type": "NPC",
        "bl_quantity": 0,
        "bl_notes": "Panagog Trial boss — Knife on the Wind. Mobile skirmisher; flanks and switches targets.",
    },
    # ── Sseskari Reed-Speaker ────────────────────────────────────────────────
    {
        "tab": "Sseskari Reed-Speaker",
        "epithet": "Voice Beneath the Stone",
        "role": "Lizardfolk Shaman — healer / controller",
        "size_type": "Medium humanoid (lizardfolk), neutral",
        "ac": 17,
        "hp": 126,
        "speed": "30 ft., swim 30 ft.",
        "prof": 4,
        "init_mod": 2,
        "cr": "9",
        "atk_per_round": 1,
        "scores": [14, 14, 18, 13, 20, 12],
        "mods": [2, 2, 4, 1, 5, 1],
        "save_prof": [False, False, True, False, True, True],
        "saves": [2, 2, 8, 1, 9, 5],
        "senses": "passive Perception 19",
        "attacks": [
            (
                "Stonekeeper's Staff",
                "Melee Weapon",
                6,
                "5 ft.",
                "1d8+3 bludgeoning plus 2d8 radiant",
                "—",
                "",
            ),
            (
                "Binding Reeds",
                "Special (Recharge 5-6)",
                "—",
                "60 ft. (20-ft. radius)",
                "—",
                "DC 17 STR",
                "Roots and spectral reeds fill a 20-foot-radius area within 60 feet until the end of "
                "Sseskari's next turn. The area is difficult terrain. Each enemy entering or starting "
                "there must succeed on the save or be restrained until the end of its turn.",
            ),
        ],
        "traits": [
            (
                "War Priest's Focus",
                "Sseskari has advantage on Constitution saves made to maintain concentration.",
            ),
            (
                "Measured Mercy",
                "When Sseskari restores hit points to a creature at 0 hit points, that creature may "
                "immediately stand without spending movement.",
            ),
            (
                "Spellcasting",
                "12th-level spellcaster. Spell save DC 17, +9 to hit with spell attacks. Requires no "
                "material components except those with a listed cost. Cantrips (at will): guidance, "
                "sacred flame, thorn whip, toll the dead. 1st (4 slots): healing word, sanctuary, shield "
                "of faith. 2nd (3 slots): lesser restoration, silence, spiritual weapon. 3rd (3 slots): "
                "dispel magic, mass healing word, spirit guardians. 4th (3 slots): banishment, freedom "
                "of movement, guardian of faith. 5th (2 slots): greater restoration, mass cure wounds. "
                "6th (1 slot): heal.",
            ),
            ("Languages", "Common, Draconic, Panagog"),
            (
                "Tactics",
                "Begins with spirit guardians only if the party closes aggressively; otherwise uses "
                "shield of faith or Binding Reeds. Uses silence against the mage only when it creates an "
                "interesting positional problem, not as a permanent shutdown. Prioritizes mass healing "
                "word to return an ally to the fight, then heal if a champion is near death. Uses "
                "banishment to create a temporary numbers advantage, usually targeting the fighter or "
                "artificer rather than the mage. Calls the team surrender when defeat is clear — never "
                "sacrifices an ally merely to prolong the trial. Demeanor: calm, observant; treats combat "
                "as sacred judgment and announces especially dangerous magic in Panagog before casting. "
                "Yields when both allies are down and cannot be restored safely, or when one ally is "
                "dying and continued combat would prevent aid — clear, formal, immediate. Team "
                "coordination: Korvath holds the center, Varka circles, Sseskari controls lanes and "
                "restores allies; they pressure but avoid eliminating one player before that character "
                "has meaningful choices. If the party struggles badly by round 3, shift the trio toward "
                "grapples, shoves, and surrender demands rather than lethal focus fire.",
            ),
        ],
        "reactions": [
            (
                "Turn the Blow (3/Day)",
                "When an ally within 30 feet takes damage, Sseskari reduces that damage by 12 (2d6 + 5).",
            ),
        ],
        "bl_type": "NPC",
        "bl_quantity": 0,
        "bl_notes": "Panagog Trial boss — Voice Beneath the Stone. Healer/controller; kill or silence to stop support.",
    },
    # ── Tharok Ghostclaw ─────────────────────────────────────────────────────
    {
        "tab": "Tharok Ghostclaw",
        "epithet": "The Thorn King (Trial I)",
        "role": "Apex Predator — solo boss",
        "size_type": "Huge monstrosity",
        "ac": 18,
        "hp": 210,
        "speed": "50 ft., climb 20 ft.",
        "prof": 5,
        "init_mod": 3,
        "cr": "14",
        "atk_per_round": 3,
        "scores": [24, 16, 19, 7, 17, 10],
        "mods": [7, 3, 4, -2, 3, 0],
        "save_prof": [True, True, True, False, True, False],
        "saves": [12, 8, 9, -2, 8, 0],
        "senses": "darkvision 120 ft., tremorsense 30 ft., passive Perception 18",
        "attacks": [
            (
                "Bite",
                "Melee Weapon",
                12,
                "10 ft.",
                "3d12+7 piercing",
                "DC 18 STR",
                "On a failed save, target is grappled and restrained (escape DC 18).",
            ),
            ("Claw", "Melee Weapon", 12, "10 ft.", "3d8+7 slashing", "—", ""),
            (
                "Tail Sweep",
                "Special",
                "—",
                "15-ft. cone",
                "2d10+7 bludgeoning",
                "DC 18 DEX",
                "Each creature in the area is knocked prone on a failed save.",
            ),
            (
                "Terrifying Roar",
                "Special (Recharge 5-6)",
                "—",
                "90 ft.",
                "—",
                "DC 17 WIS",
                "Each creature that fails is frightened for 1 minute.",
            ),
        ],
        "traits": [
            (
                "Multiattack",
                "Tharok makes one Bite attack and two Claw attacks (or replaces one Claw with a Tail Sweep).",
            ),
            (
                "Ghostscale Camouflage",
                "Invisible while motionless in natural cover. Invisibility ends when it moves or attacks. "
                "After moving through foliage, Tharok may Hide as a bonus action.",
            ),
            (
                "Apex Senses",
                "Advantage on Perception and Survival checks that rely on scent; detects bleeding "
                "creatures within 60 ft.",
            ),
            (
                "Patient Predator",
                "Tharok has advantage on initiative rolls. During the first round of combat, its "
                "movement doesn't provoke opportunity attacks from creatures that haven't yet acted.",
            ),
            (
                "Legendary Resistance (2/Day)",
                "If Tharok fails a saving throw, it can choose to succeed instead.",
            ),
            (
                "Legendary Actions (3/round)",
                "Tharok can take 3 legendary actions, choosing from the options below. Only one "
                "legendary action can be used at a time and only at the end of another creature's turn. "
                "Tharok regains spent legendary actions at the start of its turn. Stalk (move up to half "
                "speed); Detect (makes a Wisdom (Perception) check); Claw (costs 1 action, one Claw "
                "attack); Ghoststep (costs 2 actions, moves up to full speed and can Hide); Savage "
                "Reversal (costs 2 actions, makes one Bite or Tail Sweep attack).",
            ),
            (
                "Behavior & Control",
                "Natural: stalks, ambushes, uses terrain, withdraws if badly outnumbered, defends "
                "territory, avoids unnecessary risk. Controlled: magically compelled to relentlessly "
                "pursue a designated target, ignores escape opportunities, fights beyond normal limits, "
                "and abandons normal hunting instincts. Signs of Control: arcane brand beneath damaged "
                "scales, magical flashes beneath skin, unnatural fixation on one victim, refusal to "
                "retreat, obvious pain when commands are issued. Ending the Control: dispel magic "
                "(5th-level effect), remove or destroy the control focus, break the controller's "
                "concentration, or use appropriate restorative magic. Once freed, Tharok is stunned for "
                "1 round before attempting to flee.",
            ),
        ],
        "bl_type": "NPC",
        "bl_quantity": 0,
        "bl_notes": "Trial I solo boss — The Thorn King. May be magically controlled (see Traits); ambush predator.",
    },
]

# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    shutil.copy2(WORKBOOK, TMP)
    wb = openpyxl.load_workbook(TMP)

    for c in CREATURES:
        add_creature_tab(wb, c)
        row = add_battle_list_row(wb, c)
        print(f"  + {c['tab']:24s} tab created, Battle List row {row}")

    wb.save(TMP)
    shutil.copy2(TMP, WORKBOOK)
    print(f"\n✓ Added {len(CREATURES)} creatures to {WORKBOOK}")
