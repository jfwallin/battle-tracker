"""Write estimated CR values for custom campaign creatures in Combat Tracker.xlsx."""
import openpyxl

WORKBOOK = r"c:\Users\jwallin\OneDrive\dnd - elemental crown\battle tracker\combat tracker.xlsx"

CR_MAP = {
    # Bosses
    # Korrum: Huge elemental commander, AC 18, 285 HP, 3 atks +11 (3d8+7 / 3d10+7),
    #   Legendary Resistance, Siege Monster, elemental command aura.
    #   Stone Giant (CR 7) has 126 HP / 2 atks. Roughly double HP, triple attacks,
    #   +legendary resistance and command traits -> CR 13.
    "Korrum": "13",

    # Veyrath: Large elemental commander, AC 17, 240 HP, 3 atks +10 (2d8+6 / 3d8+6),
    #   Legendary Resistance, Fire Form, elemental command aura.
    #   Fire Elemental (CR 5) has 102 HP / 2 atks. More than double HP, extra attack,
    #   legendary resistance, ranged option -> CR 11.
    "Veyrath": "11",

    # Ganador soldiers
    # Legionnaire: AC 16, 30 HP, +5/1d8+3, formation traits (effectively AC 17 in formation).
    #   Hobgoblin (CR 1/2) has 11 HP / AC 18. Higher HP, similar attack, formation utility -> CR 1.
    "Ganador Legionnaire": "1",

    # Marksman: AC 14, 24 HP, +6/1d8+3 ranged, Volley Training (further bonus in melee).
    #   Hobgoblin base stats but ranged specialist with bonus to hit vs engaged targets -> CR 1/2.
    "Ganador Marksman": "1/2",

    # Veteran: AC 17, 45 HP, +6/1d10+3, Command Presence aura, Command bonus action.
    #   Hobgoblin Captain (CR 3) has 52 HP / 2 atks. Similar AC/HP, one fewer attack
    #   but command aura adds party-wide value. CR 2.
    "Ganador Veteran": "2",

    # Spellcasters
    # War Mage: AC 13, 35 HP, wizard 7-9, Fireball/Lightning Bolt, concentration advantage.
    #   MM Mage (CR 6) is 40 HP with similar spell suite. Slightly lower HP -> CR 6.
    "War Mage": "6",

    # Arcane Specialist: AC 14, 30 HP, wizard 6-8, Counterspell/Dispel focus, anti-magic role.
    #   Slightly lower level than War Mage, lower HP, but Counterspell is high-impact -> CR 5.
    "Arcane Specialist": "5",

    # War Priest: AC 15, 38 HP, cleric 5-7, healer/sustain role.
    #   MM Priest (CR 2) is much weaker. With Mass Healing Word and Revivify -> CR 4.
    "War Priest": "4",

    # Illusionist: AC 13, 32 HP, wizard illusion ~11, Greater Invisibility, Hypnotic Pattern,
    #   Hallucinatory Terrain, Nondetection passive. High-level spell slots.
    #   Archmage is CR 12; this creature is notably weaker but with high-impact illusion
    #   utility. CR 7.
    "Illusionist": "7",

    # Anchor Casters: AC 12, 18 HP, non-combatants — no attacks, ritual only.
    #   Comparable to a cultist (CR 1/8) but with elemental ward resistances.
    #   Effectively CR 1/4 each (they are objectives, not threats).
    "Fire Anchor Caster": "1/4",
    "Earth Anchor Caster": "1/4",
    "Water Anchor Caster": "1/4",
    "Air Anchor Caster": "1/4",
}

wb = openpyxl.load_workbook(WORKBOOK)
for tab, cr in CR_MAP.items():
    if tab not in wb.sheetnames:
        print(f"SKIP (not found): {tab}")
        continue
    ws = wb[tab]
    ws["A12"] = "CR"
    ws["B12"] = cr
    print(f"  {tab}: CR {cr}")
wb.save(WORKBOOK)
print("Done.")
