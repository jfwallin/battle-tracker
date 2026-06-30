"""One-shot script: write CR values into B12 of each creature tab in Combat Tracker.xlsx."""
import openpyxl

WORKBOOK = r"c:\Users\jwallin\OneDrive\dnd - elemental crown\battle tracker\combat tracker.xlsx"

# CR values per tab name. Standard MM values where known.
# Custom campaign creatures (Korrum, Veyrath, Ganador*, Anchor Casters) are left blank
# — fill those in manually or edit this dict before running.
CR_MAP = {
    # Campaign bosses / named NPCs — set these yourself
    "Korrum": "",              # custom boss
    "Veyrath": "",             # custom boss

    # Ganador soldiers — custom, estimate based on role
    "Ganador Legionnaire": "",
    "Ganador Marksman": "",
    "Ganador Veteran": "",
    "War Mage": "",
    "Arcane Specialist": "",
    "War Priest": "",
    "Illusionist": "",

    # Elemental anchor casters — custom
    "Fire Anchor Caster": "",
    "Earth Anchor Caster": "",
    "Water Anchor Caster": "",
    "Air Anchor Caster": "",

    # MM Elementals (MM p.125)
    "Fire Elemental": "5",
    "Water Elemental": "5",
    "Air Elemental": "5",
    "Earth Elemental": "5",

    # MM Mephits (MM p.215-217) — all CR 1/4
    "Magma Mephit": "1/2",
    "Steam Mephit": "1/4",
    "Ice Mephit": "1/2",
    "Dust Mephit": "1/2",
    "Mud Mephit": "1/4",
    "Smoke Mephit": "1/4",

    # MM Undead
    "Skeleton": "1/4",         # MM p.272
    "Zombie": "1/4",           # MM p.316
    "Shadow": "1/2",           # MM p.269
    "Ghoul": "1",              # MM p.148
    "Wight": "3",              # MM p.300
    "Specter": "1",            # MM p.279
    "Wraith": "5",             # MM p.302

    # MM Humanoids
    "Hobgoblin": "1/2",        # MM p.186
    "Goblin": "1/4",           # MM p.166
    "Orc": "1/2",              # MM p.246
    "Ogre": "2",               # MM p.237
    "Bugbear": "1",            # MM p.33
    "Orc War Chief": "4",      # MM p.246

    # MM Giants
    "Stone Giant": "7",        # MM p.156
    "Frost Giant": "8",        # MM p.155
    "Fire Giant": "9",         # MM p.154
    "Storm Giant": "13",       # MM p.156

    # MM Dragons (Adult)
    "Adult Red Dragon": "17",   # MM p.98
    "Adult Blue Dragon": "16",  # MM p.91
    "Adult Green Dragon": "15", # MM p.94
    "Adult Black Dragon": "14", # MM p.88
    "Adult White Dragon": "13", # MM p.101
}

wb = openpyxl.load_workbook(WORKBOOK)

updated = []
skipped = []

for tab, cr in CR_MAP.items():
    if tab not in wb.sheetnames:
        print(f"  SKIP (tab not found): {tab}")
        continue
    if not cr:
        skipped.append(tab)
        continue
    ws = wb[tab]
    ws["A12"] = "CR"
    ws["B12"] = cr
    updated.append(f"{tab} → {cr}")

wb.save(WORKBOOK)

print(f"\nWrote CR to {len(updated)} sheets:")
for u in updated:
    print(f"  {u}")

if skipped:
    print(f"\nLeft blank (custom creatures — fill in manually):")
    for s in skipped:
        print(f"  {s}")
