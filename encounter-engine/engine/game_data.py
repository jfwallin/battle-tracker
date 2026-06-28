"""Game Data (optional `Game_Data.xlsx`) — Stage 3.

Currently holds **party profiles**: reusable rosters of PCs/allies that can be attached to a
battle and imported into a live encounter, so battle definitions can stay enemies-only instead
of baking the party into each one. The app works fine without this file.

A single flat ``Party Profiles`` sheet keyed by Profile ID (one row per character) — easy to
hand-edit. Pure (no Flask) for unit testing.
"""
from __future__ import annotations
import os
from typing import Optional

import openpyxl
from openpyxl.styles import Font

from .models import Combatant, make_id

PARTY_SHEET = "Party Profiles"
PARTY_HEADER = [
    "Profile ID", "Character Name", "Type", "Default AC",
    "Default Max HP", "Initiative Modifier", "Notes",
]

VARIANTS_SHEET = "Variants"
VARIANTS_HEADER = [
    "Variant ID", "Name", "Base Source", "AC", "Max HP", "Init Mod",
    "Attacks/Round", "To-Hit", "Resistances", "Immunities", "Notes",
]
_VARIANT_COLS = [
    "id", "name", "base_source", "ac", "max_hp", "init_mod",
    "attacks_per_round", "to_hit", "resistances", "immunities", "notes",
]

RECURRING_SHEET = "Recurring NPCs"
RECURRING_HEADER = ["NPC ID", "Base Source", "Display Name", "Variant ID", "Notes", "Status"]
_RECURRING_COLS = ["id", "base_source", "name", "variant_id", "notes", "status"]


def _open(path: str):
    return openpyxl.load_workbook(path, data_only=True)


def _to_int(v) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None


def _cell(ws, r, c) -> str:
    v = ws.cell(row=r, column=c).value
    return str(v).strip() if v is not None else ""


def _header_row(ws) -> int:
    """Find the row whose column A is 'Profile ID' (else row 1)."""
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value and str(cell.value).strip().lower() in ("profile id", "profile_id"):
            return cell.row
    return 1


def _read_member_rows(ws) -> list[dict]:
    """All character rows (raw), each tagged with its profile id."""
    hr = _header_row(ws)
    members = []
    for r in range(hr + 1, ws.max_row + 1):
        pid = _cell(ws, r, 1)
        name = _cell(ws, r, 2)
        if not pid or not name:
            continue
        members.append({
            "profile_id": pid,
            "name": name,
            "type": _cell(ws, r, 3) or "PC",
            "ac": _to_int(ws.cell(row=r, column=4).value),
            "max_hp": _to_int(ws.cell(row=r, column=5).value),
            "init_mod": _to_int(ws.cell(row=r, column=6).value) or 0,
            "notes": _cell(ws, r, 7),
        })
    return members


def list_party_profiles(game_data_path: str) -> list[dict]:
    """Return profiles as ``[{id, members:[...], count}]`` (empty if no sheet)."""
    wb = _open(game_data_path)
    if PARTY_SHEET not in wb.sheetnames:
        return []
    by_id: dict[str, dict] = {}
    order: list[str] = []
    for m in _read_member_rows(wb[PARTY_SHEET]):
        pid = m["profile_id"]
        if pid not in by_id:
            by_id[pid] = {"id": pid, "members": []}
            order.append(pid)
        by_id[pid]["members"].append({k: m[k] for k in ("name", "type", "ac", "max_hp", "init_mod", "notes")})
    profiles = [by_id[p] for p in order]
    for p in profiles:
        p["count"] = len(p["members"])
    return profiles


def get_party_profile(game_data_path: str, profile_id: str) -> dict:
    profile = next((p for p in list_party_profiles(game_data_path) if p["id"] == profile_id), None)
    if profile is None:
        raise ValueError(f"Party profile '{profile_id}' not found.")
    return profile


def build_party_combatants(profile: dict) -> list[Combatant]:
    """Turn a party profile into Combatants (PCs/allies, no creature source)."""
    out = []
    for m in profile.get("members", []):
        out.append(Combatant(
            id=make_id(),
            name=m.get("name", ""),
            combatant_type=m.get("type") or "PC",
            source_tab=None,
            initiative=None,
            initiative_mod=m.get("init_mod") or 0,
            ac=m.get("ac"),
            max_hp=m.get("max_hp"),
            notes=m.get("notes", ""),
        ))
    return out


def apply_party_profile(encounter, game_data_path: str, profile_id: str) -> int:
    """Prepend a party profile's members to an encounter. Returns how many were added."""
    profile = get_party_profile(game_data_path, profile_id)
    party = build_party_combatants(profile)
    for c in party:
        encounter.combatants[c.id] = c
    # Party leads the roster (before the enemies already in order).
    encounter.order[:0] = [c.id for c in party]
    return len(party)


# ── Write side (editor) ─────────────────────────────────────────────────────────

def _member_values(profile_id: str, m: dict) -> list:
    return [
        profile_id, m.get("name", ""), m.get("type") or "PC",
        m.get("ac", ""), m.get("max_hp", ""), m.get("init_mod", 0), m.get("notes", ""),
    ]


def save_party_profile(game_data_path: str, profile: dict) -> dict:
    """Create or replace a profile's rows in Game_Data.xlsx (creating the file if needed).

    profile = {id, members:[{name,type,ac,max_hp,init_mod,notes}, ...]}.
    """
    pid = (profile.get("id") or "").strip()
    if not pid:
        raise ValueError("Profile ID is required.")

    if os.path.exists(game_data_path):
        wb = openpyxl.load_workbook(game_data_path)
        if PARTY_SHEET not in wb.sheetnames:
            wb.create_sheet(PARTY_SHEET, 0)
    else:
        wb = openpyxl.Workbook()
        wb.active.title = PARTY_SHEET
    ws = wb[PARTY_SHEET]

    # Read existing rows for OTHER profiles, then rewrite the whole sheet.
    existing = [m for m in _read_member_rows(ws) if m["profile_id"] != pid]
    for row in list(ws.iter_rows()):
        for cell in row:
            cell.value = None
    ws.append(PARTY_HEADER)
    for c in ws[1]:
        c.font = Font(bold=True)
    for m in existing:
        ws.append(_member_values(m["profile_id"], m))
    for m in profile.get("members", []):
        if not (m.get("name") or "").strip():
            continue
        ws.append(_member_values(pid, m))

    wb.save(game_data_path)
    return {"id": pid, "count": len([m for m in profile.get("members", []) if (m.get("name") or "").strip()])}


def delete_party_profile(game_data_path: str, profile_id: str) -> None:
    if not os.path.exists(game_data_path):
        raise ValueError("Game_Data.xlsx not found.")
    wb = openpyxl.load_workbook(game_data_path)
    if PARTY_SHEET not in wb.sheetnames:
        raise ValueError(f"No '{PARTY_SHEET}' sheet.")
    ws = wb[PARTY_SHEET]
    remaining = [m for m in _read_member_rows(ws) if m["profile_id"] != profile_id]
    for row in list(ws.iter_rows()):
        for cell in row:
            cell.value = None
    ws.append(PARTY_HEADER)
    for c in ws[1]:
        c.font = Font(bold=True)
    for m in remaining:
        ws.append(_member_values(m["profile_id"], m))
    wb.save(game_data_path)


# ── Variants (reusable stat-block modifications) ─────────────────────────────────

def _header_row_for(ws, label: str) -> int:
    label = label.lower()
    for row in ws.iter_rows(min_col=1, max_col=1):
        if row[0].value and str(row[0].value).strip().lower() == label:
            return row[0].row
    return 1


def _apply_expr(current, expr):
    """Apply a variant expression to a numeric stat:
    "20" = absolute, "+3"/"-2" = delta, "x1.5"/"*0.5" = multiply. Blank leaves it unchanged.
    Delta/multiply on a missing (None) value are skipped; absolute can set from None."""
    expr = str(expr or "").strip()
    if not expr:
        return current
    try:
        if expr[0] in "+-":
            return current if current is None else current + int(round(float(expr)))
        if expr[0] in "xX*":
            return current if current is None else int(round(current * float(expr.lstrip("xX*"))))
        return int(round(float(expr)))  # absolute
    except ValueError:
        return current


def _append_csv(existing: str, additions: str) -> str:
    have = [s.strip() for s in (existing or "").split(",") if s.strip()]
    low = {s.lower() for s in have}
    for a in (additions or "").split(","):
        a = a.strip()
        if a and a.lower() not in low:
            have.append(a)
            low.add(a.lower())
    return ", ".join(have)


def _read_variant_rows(ws) -> list[dict]:
    hr = _header_row_for(ws, "variant id")
    out = []
    for r in range(hr + 1, ws.max_row + 1):
        vid = _cell(ws, r, 1)
        if not vid:
            continue
        out.append({key: _cell(ws, r, i + 1) for i, key in enumerate(_VARIANT_COLS)})
    return out


def list_variants(game_data_path: str) -> list[dict]:
    wb = _open(game_data_path)
    if VARIANTS_SHEET not in wb.sheetnames:
        return []
    return _read_variant_rows(wb[VARIANTS_SHEET])


def get_variant(game_data_path: str, variant_id: str) -> dict:
    v = next((x for x in list_variants(game_data_path) if x["id"] == variant_id), None)
    if v is None:
        raise ValueError(f"Variant '{variant_id}' not found.")
    return v


def apply_variant(combatant, variant: dict) -> None:
    """Layer a variant's modifications onto a freshly-built Combatant, in place."""
    name = (variant.get("name") or "").strip()
    if name and f"({name})" not in combatant.name:
        combatant.name = f"{combatant.name} ({name})"

    if variant.get("ac"):
        combatant.ac = _apply_expr(combatant.ac, variant["ac"])
    if variant.get("max_hp"):
        new_hp = _apply_expr(combatant.max_hp, variant["max_hp"])
        combatant.max_hp = new_hp
        if combatant.is_group and new_hp is not None:
            for m in combatant.members:
                m.max_hp = new_hp
    if variant.get("init_mod"):
        combatant.initiative_mod = _apply_expr(combatant.initiative_mod, variant["init_mod"]) or 0
    if variant.get("attacks_per_round"):
        combatant.attacks_per_round = (
            _apply_expr(combatant.attacks_per_round, variant["attacks_per_round"])
            or combatant.attacks_per_round
        )
    if variant.get("to_hit"):
        for a in combatant.attacks:
            if a.to_hit is not None:
                a.to_hit = _apply_expr(a.to_hit, variant["to_hit"])
    if variant.get("resistances"):
        combatant.resistances = _append_csv(combatant.resistances, variant["resistances"])
    if variant.get("immunities"):
        combatant.immunities = _append_csv(combatant.immunities, variant["immunities"])
    if variant.get("notes"):
        combatant.notes = (f"{combatant.notes} | {variant['notes']}".strip(" |")
                           if combatant.notes else variant["notes"])


def _variant_values(v: dict) -> list:
    return [v.get(k, "") for k in _VARIANT_COLS]


def save_variant(game_data_path: str, variant: dict) -> dict:
    """Create or replace one variant row (creating Game_Data.xlsx / the sheet if needed)."""
    vid = (variant.get("id") or "").strip()
    if not vid:
        raise ValueError("Variant ID is required.")

    if os.path.exists(game_data_path):
        wb = openpyxl.load_workbook(game_data_path)
        if VARIANTS_SHEET not in wb.sheetnames:
            wb.create_sheet(VARIANTS_SHEET)
    else:
        wb = openpyxl.Workbook()
        wb.active.title = VARIANTS_SHEET
    ws = wb[VARIANTS_SHEET]

    existing = [v for v in _read_variant_rows(ws) if v["id"] != vid]
    for row in list(ws.iter_rows()):
        for cell in row:
            cell.value = None
    ws.append(VARIANTS_HEADER)
    for c in ws[1]:
        c.font = Font(bold=True)
    for v in existing:
        ws.append(_variant_values(v))
    ws.append(_variant_values({**variant, "id": vid}))
    wb.save(game_data_path)
    return {"id": vid}


def delete_variant(game_data_path: str, variant_id: str) -> None:
    if not os.path.exists(game_data_path):
        raise ValueError("Game_Data.xlsx not found.")
    wb = openpyxl.load_workbook(game_data_path)
    if VARIANTS_SHEET not in wb.sheetnames:
        raise ValueError(f"No '{VARIANTS_SHEET}' sheet.")
    ws = wb[VARIANTS_SHEET]
    remaining = [v for v in _read_variant_rows(ws) if v["id"] != variant_id]
    for row in list(ws.iter_rows()):
        for cell in row:
            cell.value = None
    ws.append(VARIANTS_HEADER)
    for c in ws[1]:
        c.font = Font(bold=True)
    for v in remaining:
        ws.append(_variant_values(v))
    wb.save(game_data_path)


# ── Recurring NPCs (named individuals based on a base creature source) ───────────

def _read_recurring_rows(ws) -> list[dict]:
    hr = _header_row_for(ws, "npc id")
    out = []
    for r in range(hr + 1, ws.max_row + 1):
        nid = _cell(ws, r, 1)
        if not nid:
            continue
        out.append({key: _cell(ws, r, i + 1) for i, key in enumerate(_RECURRING_COLS)})
    return out


def list_recurring_npcs(game_data_path: str) -> list[dict]:
    wb = _open(game_data_path)
    if RECURRING_SHEET not in wb.sheetnames:
        return []
    return _read_recurring_rows(wb[RECURRING_SHEET])


def get_recurring_npc(game_data_path: str, npc_id: str) -> dict:
    n = next((x for x in list_recurring_npcs(game_data_path) if x["id"] == npc_id), None)
    if n is None:
        raise ValueError(f"Recurring NPC '{npc_id}' not found.")
    return n


def _recurring_values(n: dict) -> list:
    return [n.get(k, "") for k in _RECURRING_COLS]


def save_recurring_npc(game_data_path: str, npc: dict) -> dict:
    nid = (npc.get("id") or "").strip()
    if not nid:
        raise ValueError("NPC ID is required.")
    if not (npc.get("base_source") or "").strip():
        raise ValueError("Base Source is required.")

    if os.path.exists(game_data_path):
        wb = openpyxl.load_workbook(game_data_path)
        if RECURRING_SHEET not in wb.sheetnames:
            wb.create_sheet(RECURRING_SHEET)
    else:
        wb = openpyxl.Workbook()
        wb.active.title = RECURRING_SHEET
    ws = wb[RECURRING_SHEET]

    existing = [n for n in _read_recurring_rows(ws) if n["id"] != nid]
    for row in list(ws.iter_rows()):
        for cell in row:
            cell.value = None
    ws.append(RECURRING_HEADER)
    for c in ws[1]:
        c.font = Font(bold=True)
    for n in existing:
        ws.append(_recurring_values(n))
    ws.append(_recurring_values({**npc, "id": nid}))
    wb.save(game_data_path)
    return {"id": nid}


def delete_recurring_npc(game_data_path: str, npc_id: str) -> None:
    if not os.path.exists(game_data_path):
        raise ValueError("Game_Data.xlsx not found.")
    wb = openpyxl.load_workbook(game_data_path)
    if RECURRING_SHEET not in wb.sheetnames:
        raise ValueError(f"No '{RECURRING_SHEET}' sheet.")
    ws = wb[RECURRING_SHEET]
    remaining = [n for n in _read_recurring_rows(ws) if n["id"] != npc_id]
    for row in list(ws.iter_rows()):
        for cell in row:
            cell.value = None
    ws.append(RECURRING_HEADER)
    for c in ws[1]:
        c.font = Font(bold=True)
    for n in remaining:
        ws.append(_recurring_values(n))
    wb.save(game_data_path)
