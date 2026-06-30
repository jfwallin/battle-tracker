"""Battle Library: reusable battle definitions that reference Combat Tracker creatures.

`Battle_Library.xlsx` has a ``Battle Index`` sheet (one row per prepared battle) plus one
definition sheet per battle, shaped exactly like the legacy ``Battle List`` so the same
row parser is reused. Creature/NPC sources in a definition resolve against the separate
``Combat Tracker.xlsx`` workbook.

Pure (no Flask) so it can be unit-tested directly.
"""
from __future__ import annotations
import os
import re
from datetime import date
from typing import Optional

import openpyxl
from openpyxl.styles import Font

from .models import Encounter
from .excel_loader import (
    ROSTER_COLUMNS,
    build_combatant_from_row,
    make_combat_tracker_resolver,
    read_battle_list_rows,
)

BATTLE_INDEX_SHEET = "Battle Index"
VALID_INIT_MODES = {"Individual", "Shared", "Grouped Attacks", "Mob"}

# 5e XP by CR (PHB/MM table)
_CR_XP: dict[str, int] = {
    "0": 10, "1/8": 25, "1/4": 50, "1/2": 100,
    "1": 200, "2": 450, "3": 700, "4": 1100, "5": 1800,
    "6": 2300, "7": 2900, "8": 3900, "9": 5000, "10": 5900,
    "11": 7200, "12": 8400, "13": 10000, "14": 11500, "15": 13000,
    "16": 15000, "17": 18000, "18": 20000, "19": 22000, "20": 25000,
    "21": 33000, "22": 41000, "23": 50000, "24": 62000, "30": 155000,
}
# 5e XP thresholds per PC per level (Easy, Medium, Hard, Deadly)
_XP_THRESH: dict[int, tuple[int, int, int, int]] = {
    1: (25, 50, 75, 100), 2: (50, 100, 150, 200), 3: (75, 150, 225, 400),
    4: (125, 250, 375, 500), 5: (250, 500, 750, 1100), 6: (300, 600, 900, 1400),
    7: (350, 750, 1100, 1700), 8: (450, 900, 1400, 2100), 9: (550, 1100, 1600, 2400),
    10: (600, 1200, 1900, 2800), 11: (800, 1600, 2400, 3600), 12: (1000, 2000, 3000, 4500),
    13: (1100, 2200, 3400, 5100), 14: (1250, 2500, 3800, 5700), 15: (1400, 2800, 4300, 6400),
    16: (1600, 3200, 4800, 7200), 17: (2000, 3900, 5900, 8800), 18: (2100, 4200, 6300, 9500),
    19: (2400, 4900, 7300, 10900), 20: (2800, 5700, 8500, 12700),
}
# Encounter XP multipliers by monster count (for encounter difficulty)
def _encounter_multiplier(count: int) -> float:
    if count == 1: return 1.0
    if count == 2: return 1.5
    if count <= 6: return 2.0
    if count <= 10: return 2.5
    if count <= 14: return 3.0
    return 4.0

def _cr_to_xp(cr: str) -> int:
    """Convert a CR string like '1/4', '5', '' to XP. Returns 0 if unknown."""
    return _CR_XP.get(str(cr).strip(), 0)

def _difficulty_tier(adjusted_xp: int, party_size: int, party_level: int) -> str:
    """Return Easy/Medium/Hard/Deadly/Unknown based on 5e thresholds."""
    level = max(1, min(20, party_level))
    thresh = _XP_THRESH.get(level, _XP_THRESH[1])
    easy, medium, hard, deadly = [t * party_size for t in thresh]
    if adjusted_xp == 0: return "Unknown"
    if adjusted_xp < easy: return "Trivial"
    if adjusted_xp < medium: return "Easy"
    if adjusted_xp < hard: return "Medium"
    if adjusted_xp < deadly: return "Hard"
    return "Deadly"

_INDEX_COLUMNS = [
    "battle_id", "name", "definition_location", "status", "tags",
    "notes", "default_party_profile", "last_modified",
]
_INDEX_HEADER = [
    "Battle ID", "Battle Name", "Definition Location", "Status",
    "Tags", "Notes", "Default Party Profile", "Last Modified",
]
# Definition-sheet header, in ROSTER_COLUMNS order (cols A–L).
_DEF_HEADER = [
    "Type", "Name", "NPC Source", "Quantity", "Group Name", "Initiative Mode",
    "HP Override", "Notes", "Starting Conditions", "Starting Status",
    "Variant ID", "Trigger",
]
_TYPE_TO_COUNT = {
    "pc": "pcs", "npc": "npcs", "ally": "allies",
    "hazard": "hazards", "event": "events",
}
_INVALID_SHEET_CHARS = r"[\\/?*\[\]:]"


def _open(path: str):
    return openpyxl.load_workbook(path, data_only=True)


def _read_index_rows(ws) -> list[dict]:
    """Read the Battle Index into row dicts (header found by 'Battle ID' in col A)."""
    header_row = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value and str(cell.value).strip().lower() in ("battle id", "battle_id"):
            header_row = cell.row
            break
    if header_row is None:
        header_row = 1

    rows: list[dict] = []
    for r in range(header_row + 1, ws.max_row + 1):
        def rv(col: int):
            v = ws.cell(row=r, column=col).value
            return str(v).strip() if v is not None else ""
        if not rv(1):
            continue  # blank Battle ID
        rows.append({key: rv(i + 1) for i, key in enumerate(_INDEX_COLUMNS)})
    return rows


def _count_and_validate(rows: list[dict], battle: dict, tracker_tabs: Optional[set]) -> None:
    """Fill battle['counts'] and append soft warnings for a definition's rows."""
    counts = battle["counts"]
    warnings = battle["warnings"]
    for row in rows:
        ctype = row.get("type", "")
        if not ctype:
            continue
        label = row.get("name", "") or row.get("group_name", "") or "?"

        qraw = row.get("quantity", "")
        try:
            qty = int(qraw) if qraw else 1
        except ValueError:
            qty = 1
            warnings.append(f"Row '{label}': quantity '{qraw}' invalid, treated as 1.")
        if qty <= 0:
            continue

        key = _TYPE_TO_COUNT.get(ctype.strip().lower())
        if key:
            counts[key] += 1
        if qty > 1:
            counts["groups"] += 1

        mode = row.get("initiative_mode", "")
        if mode and mode not in VALID_INIT_MODES:
            warnings.append(f"Row '{label}': initiative mode '{mode}' unknown; will use Individual.")

        src = row.get("npc_source", "")
        if src and tracker_tabs is not None and src not in tracker_tabs:
            warnings.append(f"Source tab '{src}' not found in Combat Tracker.")

        if row.get("trigger"):
            warnings.append(f"Row '{label}': Trigger not supported yet (Stage 5); ignored.")


def list_battles(lib_path: str, tracker_path: Optional[str] = None) -> list[dict]:
    """List prepared battles from the Battle Index with participant counts and warnings.

    Counts/validation parse each definition sheet's rows but do NOT resolve creature stats
    (cheap). If a tracker path is given, broken source-tab references are flagged by
    comparing names to the tracker's sheet names (no cell reads).
    """
    wb = _open(lib_path)
    if BATTLE_INDEX_SHEET not in wb.sheetnames:
        raise ValueError(f"Battle Library has no '{BATTLE_INDEX_SHEET}' sheet.")

    tracker_tabs: Optional[set] = None
    if tracker_path:
        try:
            tracker_tabs = set(_open(tracker_path).sheetnames)
        except Exception:
            tracker_tabs = None  # validation downgraded; not fatal

    battles: list[dict] = []
    for ir in _read_index_rows(wb[BATTLE_INDEX_SHEET]):
        battle = {
            "id": ir["battle_id"],
            "name": ir["name"] or ir["battle_id"],
            "status": ir["status"],
            "tags": [t.strip() for t in ir["tags"].split(",") if t.strip()],
            "notes": ir["notes"],
            "default_party_profile": ir["default_party_profile"],
            "definition_location": ir["definition_location"],
            "last_modified": ir["last_modified"],
            "counts": {"pcs": 0, "npcs": 0, "allies": 0, "hazards": 0, "events": 0, "groups": 0},
            "warnings": [],
        }
        loc = ir["definition_location"]
        if not loc:
            battle["warnings"].append("No definition location set.")
        elif loc not in wb.sheetnames:
            battle["warnings"].append(f"Definition sheet '{loc}' not found.")
        else:
            _count_and_validate(read_battle_list_rows(wb[loc]), battle, tracker_tabs)
        battles.append(battle)
    return battles


def load_battle(lib_path: str, tracker_path: str, battle_id: str,
                game_data_path: Optional[str] = None) -> Encounter:
    """Resolve a battle definition into a self-contained live Encounter.

    Creature stats are pulled from Combat Tracker.xlsx at load time; a row's Variant ID (if any)
    is resolved from Game_Data.xlsx and layered on. Everything is snapshotted into the Encounter,
    so later edits to source workbooks do not affect an in-progress battle.
    Raises ValueError on hard problems (unknown battle, missing definition sheet).
    """
    wb = _open(lib_path)
    if BATTLE_INDEX_SHEET not in wb.sheetnames:
        raise ValueError(f"Battle Library has no '{BATTLE_INDEX_SHEET}' sheet.")

    match = next((r for r in _read_index_rows(wb[BATTLE_INDEX_SHEET])
                  if r["battle_id"] == battle_id), None)
    if match is None:
        raise ValueError(f"Battle '{battle_id}' not found in Battle Index.")

    loc = match["definition_location"]
    if not loc:
        raise ValueError(f"Battle '{match['name'] or battle_id}' has no definition location.")
    if loc not in wb.sheetnames:
        raise ValueError(
            f"Battle '{match['name'] or battle_id}': definition sheet '{loc}' not found."
        )

    rows = read_battle_list_rows(wb[loc])
    resolve = make_combat_tracker_resolver(tracker_path)

    # Reusable variants (Game_Data.xlsx), if available.
    variants: dict = {}
    if game_data_path:
        try:
            from .game_data import list_variants
            variants = {v["id"]: v for v in list_variants(game_data_path)}
        except Exception:
            variants = {}

    # source_path points at the Combat Tracker so Excel export can still write HP/status
    # back into the creature tabs.
    enc = Encounter(source_path=tracker_path)
    for row in rows:
        combatant = build_combatant_from_row(row, resolve)
        if combatant is None:
            continue
        vid = row.get("variant_id")
        if vid and vid in variants:
            from .game_data import apply_variant
            apply_variant(combatant, variants[vid])
        enc.combatants[combatant.id] = combatant
        enc.order.append(combatant.id)
    return enc


# ── Battle preview stats (HP/CR estimates with variant support) ─────────────────

def _apply_hp_expr(current: Optional[int], expr) -> Optional[int]:
    """Apply a variant HP expression to a base HP value (mirrors game_data._apply_expr)."""
    import math
    if expr is None or expr == "":
        return current
    if isinstance(expr, (int, float)):
        return int(round(expr))
    expr = str(expr).strip()
    if not expr:
        return current
    try:
        if expr[0] in "+-":
            return current if current is None else current + int(round(float(expr)))
        if expr[0] in "xX*":
            return current if current is None else math.ceil(current * float(expr.lstrip("xX*")))
        return int(round(float(expr)))
    except ValueError:
        return current


def battle_preview_stats(
    lib_path: str,
    tracker_path: Optional[str],
    battle_id: str,
    variants: Optional[dict] = None,
    party_size: int = 4,
    party_level: int = 5,
) -> dict:
    """Compute preview stats for one battle: total HP, total XP, difficulty.

    Reads HP and CR from Combat Tracker tabs (cheap — only B8/B12 per sheet),
    applies variant HP expressions if provided. Returns:
      {total_hp, total_xp, adjusted_xp, difficulty, enemy_count, cr_list}
    """
    wb_lib = _open(lib_path)
    if BATTLE_INDEX_SHEET not in wb_lib.sheetnames:
        return {}
    match = next((r for r in _read_index_rows(wb_lib[BATTLE_INDEX_SHEET])
                  if r["battle_id"] == battle_id), None)
    if match is None:
        return {}
    loc = match["definition_location"]
    if not loc or loc not in wb_lib.sheetnames:
        return {}

    rows = read_battle_list_rows(wb_lib[loc])

    # Open tracker workbook once for HP/CR lookups.
    tracker_wb = None
    if tracker_path:
        try:
            tracker_wb = _open(tracker_path)
        except Exception:
            tracker_wb = None

    total_hp = 0
    total_xp = 0
    enemy_count = 0
    cr_list: list[str] = []
    all_cr_known = True

    for row in rows:
        ctype = (row.get("type") or "").strip().lower()
        if ctype not in ("npc", "ally"):
            continue
        qraw = row.get("quantity", "")
        try:
            qty = int(qraw) if qraw else 1
        except ValueError:
            qty = 1
        if qty <= 0:
            continue

        src = row.get("npc_source", "").strip()
        base_hp: Optional[int] = None
        cr_str = ""

        if tracker_wb and src and src in tracker_wb.sheetnames:
            ws = tracker_wb[src]
            hp_val = ws["B8"].value
            cr_val = ws["B12"].value
            try:
                base_hp = int(hp_val) if hp_val is not None else None
            except (ValueError, TypeError):
                base_hp = None
            cr_str = str(cr_val).strip() if cr_val is not None else ""

        # Apply HP override from the roster row.
        hp_override_raw = row.get("hp_override", "")
        if hp_override_raw:
            try:
                base_hp = int(hp_override_raw)
            except ValueError:
                pass

        # Apply variant HP modifier if available.
        vid = row.get("variant_id", "").strip()
        if vid and variants and vid in variants:
            v = variants[vid]
            hp_expr = v.get("max_hp", "")
            if hp_expr not in (None, ""):
                base_hp = _apply_hp_expr(base_hp, hp_expr)

        per_unit_hp = base_hp or 0
        total_hp += per_unit_hp * qty
        enemy_count += qty

        xp = _cr_to_xp(cr_str)
        total_xp += xp * qty
        for _ in range(qty):
            cr_list.append(cr_str or "?")
        if not cr_str:
            all_cr_known = False

    multiplier = _encounter_multiplier(enemy_count)
    adjusted_xp = int(total_xp * multiplier)
    difficulty = _difficulty_tier(adjusted_xp, party_size, party_level) if all_cr_known else "Unknown"

    return {
        "total_hp": total_hp,
        "total_xp": total_xp,
        "adjusted_xp": adjusted_xp,
        "multiplier": multiplier,
        "difficulty": difficulty,
        "enemy_count": enemy_count,
        "cr_list": cr_list,
    }


# ── Builder support: list sources, read a definition, write/duplicate/delete ────

def list_sources(tracker_path: str) -> list[dict]:
    """Creature/NPC sources available in Combat Tracker for the builder's picker.

    A creature tab is any sheet (other than 'Battle List') with a name in B3.
    """
    wb = _open(tracker_path)
    out: list[dict] = []
    for name in wb.sheetnames:
        if name == "Battle List":
            continue
        ws = wb[name]
        disp = ws["B3"].value
        if not disp:
            continue
        out.append({
            "tab": name,
            "name": str(disp).strip(),
            "role": str(ws["B5"].value).strip() if ws["B5"].value else "",
            "ac": ws["B7"].value,
            "max_hp": ws["B8"].value,
            "cr": str(ws["B12"].value).strip() if ws["B12"].value is not None else "",
        })
    out.sort(key=lambda s: s["name"].lower())
    return out


def get_battle_definition(lib_path: str, battle_id: str) -> dict:
    """Return a battle's metadata + roster rows for editing in the builder."""
    wb = _open(lib_path)
    if BATTLE_INDEX_SHEET not in wb.sheetnames:
        raise ValueError(f"Battle Library has no '{BATTLE_INDEX_SHEET}' sheet.")
    match = next((r for r in _read_index_rows(wb[BATTLE_INDEX_SHEET])
                  if r["battle_id"] == battle_id), None)
    if match is None:
        raise ValueError(f"Battle '{battle_id}' not found in Battle Index.")
    loc = match["definition_location"]
    rows = []
    if loc and loc in wb.sheetnames:
        rows = [r for r in read_battle_list_rows(wb[loc]) if r.get("type")]
    return {
        "id": match["battle_id"],
        "name": match["name"],
        "definition_location": loc,
        "status": match["status"],
        "tags": [t.strip() for t in match["tags"].split(",") if t.strip()],
        "notes": match["notes"],
        "default_party_profile": match["default_party_profile"],
        "rows": rows,
    }


def _next_battle_id(index_rows: list[dict]) -> str:
    nums = [int(m.group(1)) for r in index_rows
            if (m := re.match(r"BTL-(\d+)$", r["battle_id"]))]
    return f"BTL-{(max(nums) + 1) if nums else 1:03d}"


def _safe_sheet_name(name: str, wb, keep: Optional[str] = None) -> str:
    """A unique, Excel-legal sheet name derived from a battle name."""
    base = re.sub(_INVALID_SHEET_CHARS, " ", name).strip()[:28] or "Battle"
    taken = set(wb.sheetnames)
    if keep:
        taken.discard(keep)
    candidate, i = base, 2
    while candidate in taken:
        candidate = f"{base} {i}"[:31]
        i += 1
    return candidate


def _index_rownum(idx, battle_id: str) -> Optional[int]:
    for r in range(1, idx.max_row + 1):
        if str(idx.cell(row=r, column=1).value or "").strip() == battle_id:
            return r
    return None


def _ensure_index_header(idx) -> None:
    a1 = str(idx.cell(row=1, column=1).value or "").strip().lower()
    if a1 in ("battle id", "battle_id"):
        return
    has_data = any(idx.cell(row=1, column=c).value for c in range(1, len(_INDEX_HEADER) + 1))
    if has_data:
        idx.insert_rows(1)
    for col, h in enumerate(_INDEX_HEADER, start=1):
        idx.cell(row=1, column=col, value=h).font = Font(bold=True)


def _row_values(r: dict) -> list:
    qty = r.get("quantity")
    if qty in (None, ""):
        qty = 1
    out = []
    for key in ROSTER_COLUMNS:
        if key == "quantity":
            out.append(qty)
        elif key == "initiative_mode":
            out.append(r.get("initiative_mode") or "Individual")
        else:
            out.append(r.get(key, ""))
    return out


def save_battle(lib_path: str, battle: dict) -> dict:
    """Create or update a battle definition in the library (creating the file if needed).

    battle = {id?, name, status?, tags?(list|str), notes?, default_party_profile?,
              definition_location?, rows:[{type,name,npc_source,quantity,group_name,
              initiative_mode,hp_override,notes,starting_conditions,starting_status}, ...]}
    Returns {id, name, definition_location}.
    """
    name = (battle.get("name") or "").strip()
    if not name:
        raise ValueError("Battle name is required.")

    if os.path.exists(lib_path):
        wb = openpyxl.load_workbook(lib_path)
        if BATTLE_INDEX_SHEET not in wb.sheetnames:
            wb.create_sheet(BATTLE_INDEX_SHEET, 0)
    else:
        wb = openpyxl.Workbook()
        wb.active.title = BATTLE_INDEX_SHEET
    idx = wb[BATTLE_INDEX_SHEET]
    _ensure_index_header(idx)

    index_rows = _read_index_rows(idx)
    battle_id = (battle.get("id") or "").strip() or _next_battle_id(index_rows)
    existing = next((r for r in index_rows if r["battle_id"] == battle_id), None)

    loc = (battle.get("definition_location")
           or (existing["definition_location"] if existing else "")).strip()
    if not loc:
        loc = _safe_sheet_name(name, wb)

    # Replace the definition sheet contents wholesale.
    if loc in wb.sheetnames:
        del wb[loc]
    ws = wb.create_sheet(loc)
    ws.append(_DEF_HEADER)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in battle.get("rows", []):
        if not (r.get("type") or "").strip():
            continue
        ws.append(_row_values(r))

    tags = battle.get("tags", "")
    if isinstance(tags, list):
        tags = ",".join(tags)
    values = [
        battle_id, name, loc, battle.get("status") or "Draft", tags,
        battle.get("notes", ""), battle.get("default_party_profile", ""),
        date.today().isoformat(),
    ]
    rownum = _index_rownum(idx, battle_id)
    if rownum:
        for col, val in enumerate(values, start=1):
            idx.cell(row=rownum, column=col, value=val)
    else:
        idx.append(values)

    wb.save(lib_path)
    return {"id": battle_id, "name": name, "definition_location": loc}


def duplicate_battle(lib_path: str, battle_id: str) -> dict:
    """Copy a battle's setup into a new Draft battle (fresh id + sheet)."""
    defn = get_battle_definition(lib_path, battle_id)
    return save_battle(lib_path, {
        "name": f"{defn['name']} (copy)",
        "status": "Draft",
        "tags": defn["tags"],
        "notes": defn["notes"],
        "default_party_profile": defn["default_party_profile"],
        "rows": defn["rows"],
    })


def delete_battle(lib_path: str, battle_id: str) -> None:
    """Remove a battle's index row and its definition sheet."""
    wb = openpyxl.load_workbook(lib_path)
    if BATTLE_INDEX_SHEET not in wb.sheetnames:
        raise ValueError(f"Battle Library has no '{BATTLE_INDEX_SHEET}' sheet.")
    idx = wb[BATTLE_INDEX_SHEET]
    rownum = _index_rownum(idx, battle_id)
    if not rownum:
        raise ValueError(f"Battle '{battle_id}' not found in Battle Index.")
    loc = str(idx.cell(row=rownum, column=3).value or "").strip()
    idx.delete_rows(rownum, 1)
    if loc and loc in wb.sheetnames:
        del wb[loc]
    wb.save(lib_path)
