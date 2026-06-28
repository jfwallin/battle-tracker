"""Battle report export (Stage 4): a readable Scorecard + a Detailed Event Log.

Built from the combat log (now enriched with hit/miss/save outcomes). The Scorecard lays
combatants out by initiative (rows) × round (columns) with a concise summary of what each one
did that round; the Detailed Log is the chronological ledger. A ``player_safe`` mode omits the
DM Notes column and combatants still hidden from the players.

Pure (no Flask) for unit testing.
"""
from __future__ import annotations
from collections import defaultdict
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font

from .models import Encounter

_HEADER_FONT = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")

_DAMAGE_TYPES = ("damage", "adjustment_damage")
_HEAL_TYPES = ("heal", "adjustment_heal")


def _max_round(encounter: Encounter) -> int:
    rounds = [e.round for e in encounter.log] or [encounter.round]
    return max(max(rounds), encounter.round, 1)


def _cell_text(events: list) -> str:
    """Concise summary of one combatant's actions in one round (as the acting source)."""
    groups: dict[str, dict] = defaultdict(
        lambda: {"hit": 0, "crit": 0, "miss": 0, "saved": 0, "failed": 0,
                 "dmg": 0, "heal": 0, "attacks": 0, "targets": set(), "cond": []})
    for e in events:
        if e.event_type in ("condition_add", "condition_remove"):
            groups["Conditions"]["cond"].append(e.notes or e.outcome)
            continue
        name = e.attack_name or {"heal": "Healing", "save": "Spell"}.get(e.event_type, e.event_type)
        g = groups[name]
        if e.target:
            g["targets"].add(e.target)
        if e.event_type in _DAMAGE_TYPES:
            g["dmg"] += max(0, e.amount)
            if e.attacks is not None:           # group volley: use its hit/crit/attacker counts
                g["crit"] += e.crits or 0
                g["hit"] += max(0, (e.hits or 0) - (e.crits or 0))
                g["attacks"] += e.attacks
            elif e.outcome == "crit":
                g["crit"] += 1
            elif e.outcome == "failed":
                g["failed"] += 1
            elif e.outcome == "saved":
                g["saved"] += 1
            elif e.amount > 0:
                g["hit"] += 1
        elif e.event_type == "miss":
            g["miss"] += 1
        elif e.event_type == "save":
            if e.outcome == "failed":
                g["failed"] += 1
            elif e.outcome in ("saved", "immune"):
                g["saved"] += 1
        elif e.event_type in _HEAL_TYPES:
            g["heal"] += e.amount

    lines = []
    for name, g in groups.items():
        if g["cond"]:
            lines.append("; ".join(g["cond"]))
            continue
        seg = []
        if g["hit"] or g["crit"]:
            hit_seg = f"{g['hit'] + g['crit']} hit" + (f" ({g['crit']} crit)" if g["crit"] else "")
            if g["attacks"]:
                hit_seg += f" of {g['attacks']}"
            seg.append(hit_seg)
        if g["failed"]:
            seg.append(f"{g['failed']} failed")
        if g["saved"]:
            seg.append(f"{g['saved']} saved")
        if g["miss"]:
            seg.append(f"{g['miss']} miss")
        if g["dmg"]:
            seg.append(f"{g['dmg']} dmg")
        if g["heal"]:
            seg.append(f"+{g['heal']} HP")
        line = f"{name}: {', '.join(seg)}" if seg else name
        tgts = ", ".join(sorted(t for t in g["targets"] if t))
        if tgts:
            line += f" → {tgts}"
        lines.append(line)
    return "\n".join(lines)


def _build_scorecard(ws, encounter: Encounter, player_safe: bool) -> None:
    max_round = _max_round(encounter)
    ws.append(["Battle Scorecard"])
    ws["A1"].font = Font(bold=True, size=14)
    header = ["Initiative / Combatant"] + [f"Round {r}" for r in range(1, max_round + 1)] + ["Total Dmg Dealt"]
    ws.append(header)
    for c in ws[2]:
        c.font = _HEADER_FONT

    # Index log events by (source name, round).
    by_src_round: dict[tuple, list] = defaultdict(list)
    dmg_by_src: dict[str, int] = defaultdict(int)
    for e in encounter.log:
        by_src_round[(e.source, e.round)].append(e)
        if e.event_type in _DAMAGE_TYPES and e.amount > 0:
            dmg_by_src[e.source] += e.amount

    for cid in encounter.order:
        c = encounter.combatants.get(cid)
        if c is None:
            continue
        if player_safe and (c.status_override or "") in ("Hidden", "Removed"):
            continue
        row = [c.name]
        for r in range(1, max_round + 1):
            row.append(_cell_text(by_src_round.get((c.name, r), [])))
        row.append(dmg_by_src.get(c.name, 0))
        ws.append(row)

    # Widths + wrapping.
    ws.column_dimensions["A"].width = 22
    from openpyxl.utils import get_column_letter
    for col in range(2, max_round + 2):
        ws.column_dimensions[get_column_letter(col)].width = 30
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = _WRAP


def _build_detailed_log(ws, encounter: Encounter, player_safe: bool) -> None:
    cols = ["Round", "Turn", "Source", "Target", "Type", "Outcome", "Roll",
            "Hits", "Attacks", "Amount", "Dmg Type", "Attack/Spell"]
    if not player_safe:
        cols.append("Notes")
    cols.append("Timestamp")
    ws.append(cols)
    for c in ws[1]:
        c.font = _HEADER_FONT
    for e in encounter.log:
        hits = "" if e.hits is None else (f"{e.hits}" + (f" ({e.crits} crit)" if e.crits else ""))
        row = [e.round, e.turn, e.source, e.target, e.event_type, e.outcome,
               e.roll if e.roll is not None else "",
               hits, e.attacks if e.attacks is not None else "",
               e.amount if e.event_type in _DAMAGE_TYPES + _HEAL_TYPES else "",
               e.damage_type, e.attack_name]
        if not player_safe:
            row.append(e.notes)
        row.append(e.ts)
        ws.append(row)
    widths = {"C": 16, "D": 16, "E": 14, "L": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def export_report(encounter: Encounter, out_path: str, player_safe: bool = False) -> str:
    """Write a battle report workbook (Scorecard + Detailed Log) to out_path. Returns out_path."""
    wb = openpyxl.Workbook()
    _build_scorecard(wb.active, encounter, player_safe)
    wb.active.title = "Scorecard"
    _build_detailed_log(wb.create_sheet("Detailed Log"), encounter, player_safe)
    wb.save(out_path)
    return out_path
