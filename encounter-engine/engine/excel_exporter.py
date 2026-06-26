"""Export Encounter state back to an Excel workbook."""
from __future__ import annotations
import shutil
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill

from .models import Encounter
from .combat import current_hp, derive_status


def export_encounter(encounter: Encounter, out_path: str) -> None:
    """
    Copy the source workbook, update HP/status in creature sheets,
    and append a Log sheet. Never overwrites the source unless out_path == source_path.
    """
    src = encounter.source_path
    if src and src != out_path:
        shutil.copy2(src, out_path)

    wb = openpyxl.load_workbook(out_path)

    # Write a Log sheet
    if "Combat Log" in wb.sheetnames:
        del wb["Combat Log"]
    log_ws = wb.create_sheet("Combat Log")

    headers = ["ID","Round","Turn","Source","Target","Type","Amount","Dmg Type","Attack","Notes","Timestamp"]
    log_ws.append(headers)
    for cell in log_ws[1]:
        cell.font = Font(bold=True)

    for e in encounter.log:
        log_ws.append([
            e.id, e.round, e.turn, e.source, e.target,
            e.event_type, e.amount if e.event_type in ("damage","heal") else "",
            e.damage_type, e.attack_name, e.notes, e.ts,
        ])

    # Update current HP and status in creature sheets
    for cid, c in encounter.combatants.items():
        if not c.source_tab or c.source_tab not in wb.sheetnames:
            continue
        ws = wb[c.source_tab]
        hp = current_hp(c, encounter.log)
        status = derive_status(c, encounter.log)
        # Write current HP to K3, status to K4 (non-destructive, out-of-formula area)
        ws["K3"] = hp
        ws["K4"] = status
        ws["K3"].font = Font(bold=True)

    # Summary sheet
    if "Encounter Summary" in wb.sheetnames:
        del wb["Encounter Summary"]
    summ = wb.create_sheet("Encounter Summary", 0)
    summ.append(["Encounter Summary", f"Exported {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    summ.append(["Round", encounter.round])
    summ.append([])
    summ.append(["Name","Type","AC","Max HP","Current HP","Status","Conditions"])
    for cid in encounter.order:
        c = encounter.combatants[cid]
        hp = current_hp(c, encounter.log)
        status = derive_status(c, encounter.log)
        summ.append([
            c.name, c.combatant_type, c.ac or "", c.max_hp or "",
            hp, status, ", ".join(c.conditions),
        ])

    wb.save(out_path)
