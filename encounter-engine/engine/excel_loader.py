"""Parse an Excel workbook into an Encounter model."""
from __future__ import annotations
import re
from typing import Callable, Optional
import openpyxl

from .models import (
    Attack, Combatant, DamageDie, Encounter, LimitedUse, Member, make_id
)

ABILITY_KEYS = ["STR", "DEX", "CON", "INT", "WIS", "CHA"]

# Section header anchors in creature sheets (uppercase, stripped)
SECTION_HEADERS = {
    "ATTACKS", "TRAITS", "ACTIONS", "BONUS ACTIONS", "REACTIONS", "BLOODIED"
}


def _cell(ws, cell_ref: str):
    """Return stripped string value of a cell, or '' if None."""
    v = ws[cell_ref].value
    return str(v).strip() if v is not None else ""


def _int_cell(ws, cell_ref: str) -> Optional[int]:
    v = ws[cell_ref].value
    if v is None:
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None


def parse_to_hit(s: str) -> Optional[int]:
    """'+10' / '10' / '—' / '' -> int or None."""
    s = s.strip().replace("—", "").replace("–", "")
    if not s:
        return None
    s = s.lstrip("+")
    try:
        return int(s)
    except ValueError:
        return None


def parse_damage(s: str) -> list[DamageDie]:
    """
    Parse strings like '2d8+6 fire' or '1d8+3 slashing plus 2d6 fire'.
    Returns list of DamageDie.
    """
    if not s or s.strip() in ("—", "–", ""):
        return []
    clauses = re.split(r"\bplus\b", s, flags=re.IGNORECASE)
    result = []
    pattern = re.compile(
        r"(\d+)\s*d\s*(\d+)\s*([+-]\s*\d+)?\s*([a-zA-Z]*)",
        re.IGNORECASE,
    )
    for clause in clauses:
        m = pattern.search(clause.strip())
        if m:
            n = int(m.group(1))
            die = int(m.group(2))
            bonus_str = (m.group(3) or "0").replace(" ", "")
            bonus = int(bonus_str) if bonus_str else 0
            dmg_type = m.group(4).strip().lower() if m.group(4) else ""
            result.append(DamageDie(n=n, die=die, bonus=bonus, damage_type=dmg_type))
    return result


def parse_save(s: str) -> tuple[Optional[int], str]:
    """'DC 17 STR' -> (17, 'STR'); '—' -> (None, '')."""
    if not s or s.strip() in ("—", "–", ""):
        return None, ""
    m = re.search(r"DC\s*(\d+)\s*([A-Z]{3})", s.strip(), re.IGNORECASE)
    if m:
        return int(m.group(1)), m.group(2).upper()
    return None, ""


def _read_section(ws, start_row: int, name_col: int = 1, desc_col: int = 2) -> list[dict]:
    """Read name/desc pairs downward from start_row until name cell is blank."""
    items = []
    row = start_row
    max_row = ws.max_row
    while row <= max_row:
        name_val = ws.cell(row=row, column=name_col).value
        if name_val is None or str(name_val).strip() == "":
            break
        name_str = str(name_val).strip()
        # Stop if we hit another section header
        if name_str.upper() in SECTION_HEADERS:
            break
        desc_val = ws.cell(row=row, column=desc_col).value
        items.append({
            "name": name_str,
            "desc": str(desc_val).strip() if desc_val is not None else "",
        })
        row += 1
    return items


def _read_attacks(ws, start_row: int) -> list[Attack]:
    """Read attack rows from start_row downward until name col (A) is blank."""
    attacks = []
    row = start_row
    max_row = ws.max_row
    while row <= max_row:
        name_val = ws.cell(row=row, column=1).value
        if name_val is None or str(name_val).strip() == "":
            break
        name_str = str(name_val).strip()
        if name_str.upper() in SECTION_HEADERS:
            break
        atk_type = str(ws.cell(row=row, column=2).value or "").strip()
        to_hit_raw = str(ws.cell(row=row, column=3).value or "").strip()
        reach_raw = str(ws.cell(row=row, column=4).value or "").strip()
        dmg_raw = str(ws.cell(row=row, column=5).value or "").strip()
        save_raw = str(ws.cell(row=row, column=6).value or "").strip()
        effect_raw = str(ws.cell(row=row, column=7).value or "").strip()

        save_dc, save_ability = parse_save(save_raw)
        attacks.append(Attack(
            name=name_str,
            attack_type=atk_type,
            to_hit=parse_to_hit(to_hit_raw),
            reach=reach_raw,
            damage_dice=parse_damage(dmg_raw),
            save_dc=save_dc,
            save_ability=save_ability,
            effect=effect_raw,
        ))
        row += 1
    return attacks


def _find_section_row(ws, header: str) -> Optional[int]:
    """Scan column A for a cell whose stripped upper value matches header."""
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value and str(cell.value).strip().upper() == header.upper():
            return cell.row
    return None


def load_creature_sheet(wb: openpyxl.Workbook, tab_name: str) -> dict:
    """Parse a creature sheet tab into a dict of stat block fields."""
    if tab_name not in wb.sheetnames:
        return {}
    ws = wb[tab_name]

    def c(ref):
        return _cell(ws, ref)

    def ci(ref):
        return _int_cell(ws, ref)

    stats: dict = {
        "name": c("B3"),
        "epithet": c("B4"),
        "role": c("B5"),
        "size_type_alignment": c("B6"),
        "ac": ci("B7"),
        "max_hp": ci("B8"),
        "speed": c("B9"),
        "proficiency_bonus": ci("B10") or 0,
        "initiative_mod": ci("B11") or 0,
        "attacks_per_round": ci("K2") or 1,
    }

    # Ability scores B15:G15, mods B16:G16, saves B18:G18
    scores = {}
    mods = {}
    saves = {}
    for i, key in enumerate(ABILITY_KEYS):
        col = 2 + i  # B=2, C=3, ...
        sv = ws.cell(row=15, column=col).value
        mv = ws.cell(row=16, column=col).value
        savev = ws.cell(row=18, column=col).value
        scores[key] = int(sv) if sv is not None else 0
        mods[key] = int(mv) if mv is not None else 0
        saves[key] = int(savev) if savev is not None else 0
    stats["ability_scores"] = scores
    stats["ability_mods"] = mods
    stats["save_bonuses"] = saves

    stats["resistances"] = c("B21")
    stats["immunities"] = c("B22")
    stats["condition_immunities"] = c("B23")
    stats["senses"] = c("B24")

    # Attacks section
    atk_row = _find_section_row(ws, "ATTACKS")
    stats["attacks"] = _read_attacks(ws, atk_row + 1) if atk_row else []

    # Traits
    traits_row = _find_section_row(ws, "TRAITS")
    stats["traits"] = _read_section(ws, traits_row + 1) if traits_row else []

    # Actions
    actions_row = _find_section_row(ws, "ACTIONS")
    stats["actions"] = _read_section(ws, actions_row + 1) if actions_row else []

    # Bonus actions
    ba_row = _find_section_row(ws, "BONUS ACTIONS")
    stats["bonus_actions"] = _read_section(ws, ba_row + 1) if ba_row else []

    # Reactions
    react_row = _find_section_row(ws, "REACTIONS")
    stats["reactions"] = _read_section(ws, react_row + 1) if react_row else []

    # Bloodied
    blood_row = _find_section_row(ws, "BLOODIED")
    stats["bloodied_effects"] = _read_section(ws, blood_row + 1) if blood_row else []

    return stats


# ── Roster parsing (shared by single-workbook load and battle-library load) ─────

# Optional columns I–L exist only on Battle Library definition sheets; legacy
# "Battle List" sheets lack them and read as "".
ROSTER_COLUMNS = [
    "type", "name", "npc_source", "quantity", "group_name", "initiative_mode",
    "hp_override", "notes", "starting_conditions", "starting_status",
    "variant_id", "trigger",
]


def read_battle_list_rows(ws) -> list[dict]:
    """Read roster rows from a Battle-List-shaped sheet into raw string dicts.

    Finds the header row by scanning column A for 'Type' (else row 1), then reads
    each subsequent row's columns A–L. Blank/trailing rows are included verbatim;
    ``build_combatant_from_row`` decides which to skip. The same column layout is
    used by the legacy "Battle List" sheet and by Battle Library definition sheets.
    """
    header_row = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value and str(cell.value).strip().lower() == "type":
            header_row = cell.row
            break
    if header_row is None:
        header_row = 1

    rows: list[dict] = []
    for r in range(header_row + 1, ws.max_row + 1):
        def rv(col: int):
            v = ws.cell(row=r, column=col).value
            return str(v).strip() if v is not None else ""
        rows.append({key: rv(i + 1) for i, key in enumerate(ROSTER_COLUMNS)})
    return rows


def build_combatant_from_row(
    row: dict, resolve: Callable[[str], dict]
) -> Optional[Combatant]:
    """Turn one raw roster-row dict into a Combatant.

    Creature stats come from ``resolve(npc_source) -> stats dict``; the resolver may
    target a different workbook (e.g. Combat Tracker.xlsx) than the roster's source.
    Returns None for rows that should be skipped (blank Type, quantity <= 0).
    """
    combatant_type = row.get("type", "")
    if not combatant_type:
        return None  # blank row

    name = row.get("name", "")
    npc_source = row.get("npc_source", "")
    quantity_raw = row.get("quantity", "")
    group_name = row.get("group_name", "")
    initiative_mode = row.get("initiative_mode", "") or "Individual"
    hp_override_raw = row.get("hp_override", "")
    notes = row.get("notes", "")

    try:
        quantity = int(quantity_raw) if quantity_raw else 1
    except ValueError:
        quantity = 1

    # Quantity 0 (or negative) means "leave this one out of the encounter".
    # Blank quantity still defaults to 1 above.
    if quantity <= 0:
        return None

    hp_override = None
    try:
        hp_override = int(hp_override_raw) if hp_override_raw else None
    except ValueError:
        pass

    # Resolve creature stats if an NPC source is specified.
    creature = resolve(npc_source) if npc_source else {}

    max_hp = hp_override if hp_override is not None else creature.get("max_hp")
    display_name = group_name if group_name else name

    is_group = quantity > 1

    members: list[Member] = []
    if is_group:
        for i in range(1, quantity + 1):
            members.append(Member(
                id=make_id(),
                name=f"{name} {i}",
                max_hp=max_hp or 0,
            ))

    cid = make_id()
    combatant = Combatant(
        id=cid,
        name=display_name,
        combatant_type=combatant_type,
        source_tab=npc_source or None,
        initiative=None,
        initiative_mod=creature.get("initiative_mod", 0),
        ac=creature.get("ac"),
        max_hp=max_hp,
        is_group=is_group,
        members=members,
        initiative_mode=initiative_mode,
        attacks_per_round=creature.get("attacks_per_round", 1),
        notes=notes,
        speed=creature.get("speed", ""),
        size_type_alignment=creature.get("size_type_alignment", ""),
        proficiency_bonus=creature.get("proficiency_bonus", 0),
        ability_scores=creature.get("ability_scores", {}),
        ability_mods=creature.get("ability_mods", {}),
        save_bonuses=creature.get("save_bonuses", {}),
        resistances=creature.get("resistances", ""),
        immunities=creature.get("immunities", ""),
        condition_immunities=creature.get("condition_immunities", ""),
        senses=creature.get("senses", ""),
        attacks=creature.get("attacks", []),
        traits=creature.get("traits", []),
        actions=creature.get("actions", []),
        bonus_actions=creature.get("bonus_actions", []),
        reactions=creature.get("reactions", []),
        bloodied_effects=creature.get("bloodied_effects", []),
    )

    # Optional battle-definition extras (cols J/K). Legacy Battle List rows lack
    # these (read as ""), so they are no-ops there.
    starting_conditions = row.get("starting_conditions", "")
    if starting_conditions:
        combatant.conditions = [
            c.strip() for c in starting_conditions.split(",") if c.strip()
        ]
    starting_status = row.get("starting_status", "")
    if starting_status:
        combatant.status_override = starting_status

    return combatant


def make_workbook_resolver(wb) -> Callable[[str], dict]:
    """Return a creature resolver `tab_name -> stats dict` bound to an open workbook."""
    return lambda tab_name: load_creature_sheet(wb, tab_name)


def make_combat_tracker_resolver(combat_tracker_path: str) -> Callable[[str], dict]:
    """Open the Combat Tracker workbook (read-only) and return a creature resolver.

    Used to resolve creature/NPC sources referenced by a battle definition that lives
    in a *different* workbook (Battle_Library.xlsx).
    """
    wb = openpyxl.load_workbook(combat_tracker_path, data_only=True)
    return make_workbook_resolver(wb)


def load_encounter(path: str) -> Encounter:
    """Parse a single workbook (its 'Battle List' sheet + creature tabs) into an Encounter."""
    wb = openpyxl.load_workbook(path, data_only=True)

    if "Battle List" not in wb.sheetnames:
        raise ValueError("Workbook has no 'Battle List' sheet")

    rows = read_battle_list_rows(wb["Battle List"])
    resolve = make_workbook_resolver(wb)

    enc = Encounter(source_path=path)
    for row in rows:
        combatant = build_combatant_from_row(row, resolve)
        if combatant is None:
            continue
        enc.combatants[combatant.id] = combatant
        enc.order.append(combatant.id)
    return enc
