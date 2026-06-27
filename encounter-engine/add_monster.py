"""
add_monster.py — helper for adding a creature to combat tracker.xlsx

Usage: edit the CREATURE dict below, then run:
  python3 add_monster.py

The script creates a new creature tab and appends a row to Battle List.
It will overwrite an existing tab with the same name if you re-run.

Cell layout matches excel_loader.py fixed-cell reads exactly.
"""
import openpyxl
import os

WORKBOOK = os.path.join(os.path.dirname(__file__), "..", "combat tracker.xlsx")

# ── Edit this dict to describe your creature ──────────────────────────────────
CREATURE = {
    # --- Identity ---
    "tab_name":   "Hobgoblin",          # must match NPC Source in Battle List
    "epithet":    "",
    "role":       "Soldier",
    "size_type":  "Medium humanoid (goblinoid), lawful evil",

    # --- Core stats ---
    "ac":         18,
    "max_hp":     11,
    "speed":      "30 ft.",
    "prof":       2,
    "init_mod":   1,                    # usually DEX modifier
    "atk_per_round": 1,                 # K2

    # --- Ability scores [STR, DEX, CON, INT, WIS, CHA] ---
    "scores": [13, 12, 12, 10, 10,  9],
    "mods":   [ 1,  1,  1,  0,  0, -1],
    "saves":  [ 1,  1,  1,  0,  0, -1],  # full bonus (not just proficiency)
    # "save_prof": [True, False, False, False, False, False],  # optional

    # --- Defenses ---
    "resistances":   "",                # e.g. "fire, poison"
    "immunities":    "",
    "cond_immune":   "",
    "senses":        "darkvision 60 ft., passive Perception 10",

    # --- Attacks: list of (name, type, to_hit, reach, damage, save, effect) ---
    # to_hit: integer or "—" for no attack roll
    # damage: "2d6+3 slashing" or "—"
    # save:   "DC 14 DEX" or "—"
    "attacks": [
        ("Longsword", "Melee Weapon", 3, "5 ft.", "1d8+1 slashing", "—", ""),
    ],

    # --- Traits: list of (name, description) ---
    "traits": [
        ("Martial Advantage",
         "Once per turn, the hobgoblin can deal an extra 2d6 damage to a creature "
         "it hits with a weapon attack if that creature is within 5 ft. of an ally "
         "of the hobgoblin that isn't incapacitated."),
    ],

    # --- Actions (beyond attacks): list of (name, description) ---
    "actions": [],

    # --- Bonus Actions: list of (name, description) ---
    "bonus_actions": [],

    # --- Reactions: list of (name, description) ---
    "reactions": [],

    # --- Bloodied effects: list of (name, description) ---
    "bloodied": [],

    # ── Battle List row ───────────────────────────────────────────────────────
    "bl_type":       "NPC",             # NPC / PC / Ally / Hazard / Event
    "bl_quantity":   1,
    "bl_group_name": "",                # shown instead of name for groups
    "bl_init_mode":  "Individual",      # Individual / Shared / Grouped Attacks / Mob
    "bl_hp_override": None,
    "bl_notes":      "",
}
# ─────────────────────────────────────────────────────────────────────────────


def write_section(ws, start_row, anchor, entries, col_b_label="Description"):
    """Write an anchor + optional header + name/desc pairs. Returns next row."""
    ws.cell(row=start_row, column=1, value=anchor)
    ws.cell(row=start_row + 1, column=1, value="Name")
    ws.cell(row=start_row + 1, column=2, value=col_b_label)
    r = start_row + 2
    for name, desc in entries:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=desc)
        r += 1
    # blank separator
    return r + 1


def add_creature(wb, c):
    tab = c["tab_name"]

    # Remove old tab if present
    if tab in wb.sheetnames:
        del wb[tab]
    ws = wb.create_sheet(tab)

    # Header / hooks
    ws["A1"] = "CREATURE STAT BLOCK"
    ws["K2"] = c["atk_per_round"]

    # Identity block (rows 3–11)
    ws["A3"] = "Name";                    ws["B3"] = tab
    ws["A4"] = "Epithet";                 ws["B4"] = c.get("epithet", "")
    ws["A5"] = "Role";                    ws["B5"] = c.get("role", "")
    ws["A6"] = "Size / Type / Alignment"; ws["B6"] = c["size_type"]
    ws["A7"] = "AC";                      ws["B7"] = c["ac"]
    ws["A8"] = "Base HP";                 ws["B8"] = c["max_hp"]
    ws["A9"] = "Speed";                   ws["B9"] = c["speed"]
    ws["A10"] = "Proficiency Bonus";      ws["B10"] = c["prof"]
    ws["A11"] = "Initiative Mod";         ws["B11"] = c["init_mod"]

    # Ability scores (rows 13–18)
    ws["A13"] = "ABILITY SCORES"
    for i, label in enumerate(["STR", "DEX", "CON", "INT", "WIS", "CHA"]):
        ws.cell(row=14, column=2 + i, value=label)
    ws["A15"] = "Score";          ws["A16"] = "Modifier"
    ws["A17"] = "Save Proficient"; ws["A18"] = "Save Bonus"
    for i, (sc, mo, sv) in enumerate(zip(c["scores"], c["mods"], c["saves"])):
        ws.cell(row=15, column=2 + i, value=sc)
        ws.cell(row=16, column=2 + i, value=mo)
        ws.cell(row=18, column=2 + i, value=sv)

    # Defenses (rows 20–24)
    ws["A20"] = "DEFENSES & SENSES"
    ws["A21"] = "Resistances";          ws["B21"] = c.get("resistances", "")
    ws["A22"] = "Immunities";           ws["B22"] = c.get("immunities", "")
    ws["A23"] = "Condition Immunities"; ws["B23"] = c.get("cond_immune", "")
    ws["A24"] = "Senses";               ws["B24"] = c.get("senses", "")

    # Attacks section (row 26)
    ws["A26"] = "ATTACKS"
    for col, label in enumerate(
        ["Name", "Type", "To Hit", "Reach / Range", "Damage", "Save", "Effect"], start=1
    ):
        ws.cell(row=27, column=col, value=label)

    r = 28
    for name, typ, hit, reach, dmg, save, eff in c["attacks"]:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=typ)
        ws.cell(row=r, column=3, value=hit)
        ws.cell(row=r, column=4, value=reach)
        ws.cell(row=r, column=5, value=dmg)
        ws.cell(row=r, column=6, value=save)
        ws.cell(row=r, column=7, value=eff)
        r += 1

    next_row = r + 1   # blank separator after attacks

    # Variable sections
    for anchor, entries in [
        ("TRAITS",        c.get("traits", [])),
        ("ACTIONS",       c.get("actions", [])),
        ("BONUS ACTIONS", c.get("bonus_actions", [])),
        ("REACTIONS",     c.get("reactions", [])),
        ("BLOODIED",      c.get("bloodied", [])),
    ]:
        if entries:
            next_row = write_section(ws, next_row, anchor, entries)

    return ws


def add_battle_list_row(wb, c):
    bl = wb["Battle List"]
    # Find first empty row in col A after the data
    insert_row = bl.max_row + 1
    for r in range(1, bl.max_row + 2):
        if bl.cell(row=r, column=1).value is None:
            insert_row = r
            break

    bl.cell(row=insert_row, column=1, value=c["bl_type"])
    bl.cell(row=insert_row, column=2, value=c["tab_name"])
    bl.cell(row=insert_row, column=3, value=c["tab_name"])
    bl.cell(row=insert_row, column=4, value=c["bl_quantity"])
    bl.cell(row=insert_row, column=5, value=c.get("bl_group_name") or None)
    bl.cell(row=insert_row, column=6, value=c["bl_init_mode"])
    bl.cell(row=insert_row, column=7, value=c.get("bl_hp_override") or None)
    bl.cell(row=insert_row, column=8, value=c.get("bl_notes") or None)
    return insert_row


if __name__ == "__main__":
    wb = openpyxl.load_workbook(WORKBOOK)
    add_creature(wb, CREATURE)
    row = add_battle_list_row(wb, CREATURE)
    wb.save(WORKBOOK)
    print(f"✓ Added '{CREATURE['tab_name']}' tab and Battle List row {row} to {WORKBOOK}")
