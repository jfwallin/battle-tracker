"""Tests for the battle report export (engine/battle_report.py)."""
import os
import sys

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from engine.battle_report import export_report
from engine.models import Encounter, Combatant
from engine.combat import apply_damage, apply_heal, log_event, add_condition


def _encounter():
    enc = Encounter(source_path="x")
    vey = Combatant(id="m", name="Veyrath", combatant_type="NPC", source_tab=None,
                    initiative=20, initiative_mod=0, ac=17, max_hp=240)
    rex = Combatant(id="r", name="Rex", combatant_type="PC", source_tab=None,
                    initiative=15, initiative_mod=0, ac=None, max_hp=None)
    enc.combatants = {"m": vey, "r": rex}
    enc.order = ["m", "r"]
    enc.round = 1
    apply_damage(enc, "r", 22, "fire", "Fireball", "Veyrath", outcome="failed", roll=9)
    log_event(enc, "miss", "m", source="Rex", attack_name="Longsword", roll=11, outcome="miss")
    enc.round = 2
    apply_damage(enc, "m", 17, "slashing", "Longsword", "Rex", outcome="hit", roll=19)
    apply_heal(enc, "r", 8, source="Rex")
    add_condition(enc, "m", "Frightened", "1 min")
    return enc


def test_report_has_both_sheets(tmp_path):
    out = str(tmp_path / "report.xlsx")
    export_report(_encounter(), out)
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Scorecard", "Detailed Log"]


def test_scorecard_layout_and_summaries(tmp_path):
    out = str(tmp_path / "report.xlsx")
    export_report(_encounter(), out)
    sc = openpyxl.load_workbook(out)["Scorecard"]
    header = [c.value for c in sc[2]]
    assert header[0] == "Initiative / Combatant"
    assert "Round 1" in header and "Round 2" in header and header[-1] == "Total Dmg Dealt"

    rows = {r[0].value: r for r in sc.iter_rows(min_row=3)}
    vey = rows["Veyrath"]
    r1 = vey[1].value  # Round 1 cell
    assert "Fireball" in r1 and "failed" in r1 and "22 dmg" in r1
    assert vey[-1].value == 22  # total damage dealt (round 1 only; longsword 17 is Rex)
    rex = rows["Rex"]
    assert "miss" in rex[1].value.lower()       # Rex missed in round 1
    assert "Longsword" in rex[2].value and "17" in rex[2].value  # hit in round 2


def test_detailed_log_columns_and_rows(tmp_path):
    out = str(tmp_path / "report.xlsx")
    export_report(_encounter(), out)
    log = openpyxl.load_workbook(out)["Detailed Log"]
    cols = [c.value for c in log[1]]
    assert cols[:6] == ["Round", "Turn", "Source", "Target", "Type", "Outcome"]
    assert "Notes" in cols
    # The miss row is present with outcome 'miss'.
    types = [r[4] for r in log.iter_rows(min_row=2, values_only=True)]
    outcomes = [r[5] for r in log.iter_rows(min_row=2, values_only=True)]
    assert "miss" in types and "miss" in outcomes
    assert "failed" in outcomes and "hit" in outcomes


def test_player_safe_drops_notes(tmp_path):
    out = str(tmp_path / "safe.xlsx")
    export_report(_encounter(), out, player_safe=True)
    cols = [c.value for c in openpyxl.load_workbook(out)["Detailed Log"][1]]
    assert "Notes" not in cols


def test_group_volley_hit_count(tmp_path):
    """A group volley applied as one damage event reports its real hit/attacker counts."""
    from engine.combat import apply_damage
    enc = Encounter(source_path="x")
    grp = Combatant(id="g", name="Hobgoblins", combatant_type="NPC", source_tab=None,
                    initiative=10, initiative_mod=0, ac=18, max_hp=88, is_group=True)
    rex = Combatant(id="r", name="Rex", combatant_type="PC", source_tab=None,
                    initiative=12, initiative_mod=0, ac=14, max_hp=40)
    enc.combatants = {"g": grp, "r": rex}
    enc.order = ["g", "r"]
    enc.round = 1
    # One pooled damage event carrying the volley summary (5 of 8 hit, 1 crit).
    apply_damage(enc, "r", 37, "piercing", "Longbow", "Hobgoblins",
                 hits=5, crits=1, attacks=8)

    out = str(tmp_path / "report.xlsx")
    export_report(enc, out)
    sc = openpyxl.load_workbook(out)["Scorecard"]
    cell = [r for r in sc.iter_rows(min_row=3, values_only=True) if r[0] == "Hobgoblins"][0][1]
    assert "5 hit" in cell and "(1 crit)" in cell and "of 8" in cell and "37 dmg" in cell

    log = openpyxl.load_workbook(out)["Detailed Log"]
    hdr = [x.value for x in log[1]]
    assert "Hits" in hdr and "Attacks" in hdr
    row = [r for r in log.iter_rows(min_row=2, values_only=True) if r[4] == "damage"][0]
    assert row[hdr.index("Attacks")] == 8
