"""
add_npcs.py — Add all Ganador military + Elemental Cult Defender NPC classes
             to combat tracker.xlsx

BEFORE RUNNING: Restore combat tracker.xlsx from OneDrive version history.
  1. Right-click combat tracker.xlsx in File Explorer
  2. Properties → Previous Versions  (or OneDrive → Version History)
  3. Restore a version from before today

THEN RUN (from battle tracker folder):
  python3 add_npcs.py

Uses safe write: modifies a /tmp copy, then replaces the original.
Existing tabs with matching names are replaced; all other data preserved.
"""
import openpyxl
import shutil
import os
import sys

WORKBOOK = "/sessions/bold-zen-pasteur/mnt/dnd - elemental crown/battle tracker/combat tracker.xlsx"
TMP      = "/tmp/combat_tracker_modified.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def write_section(ws, start_row, anchor, entries, col_b_label="Description"):
    ws.cell(row=start_row,   column=1, value=anchor)
    ws.cell(row=start_row+1, column=1, value="Name")
    ws.cell(row=start_row+1, column=2, value=col_b_label)
    r = start_row + 2
    for name, desc in entries:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=desc)
        r += 1
    return r + 1   # blank separator row

def add_creature_tab(wb, c):
    tab = c["tab"]
    if tab in wb.sheetnames:
        del wb[tab]
    ws = wb.create_sheet(tab)

    ws["A1"] = "CREATURE STAT BLOCK"
    ws["K2"] = c.get("atk_per_round", 1)

    ws["A3"]  = "Name";                    ws["B3"]  = tab
    ws["A4"]  = "Epithet";                 ws["B4"]  = c.get("epithet", "")
    ws["A5"]  = "Role";                    ws["B5"]  = c.get("role", "")
    ws["A6"]  = "Size / Type / Alignment"; ws["B6"]  = c.get("size_type", "Medium humanoid, neutral")
    ws["A7"]  = "AC";                      ws["B7"]  = c["ac"]
    ws["A8"]  = "Base HP";                 ws["B8"]  = c["hp"]
    ws["A9"]  = "Speed";                   ws["B9"]  = c.get("speed", "30 ft.")
    ws["A10"] = "Proficiency Bonus";       ws["B10"] = c.get("prof", 2)
    ws["A11"] = "Initiative Mod";          ws["B11"] = c.get("init_mod", 0)

    # Ability scores
    ws["A13"] = "ABILITY SCORES"
    for i, lbl in enumerate(["STR","DEX","CON","INT","WIS","CHA"]):
        ws.cell(row=14, column=2+i, value=lbl)
    ws["A15"] = "Score";            ws["A16"] = "Modifier"
    ws["A17"] = "Save Proficient";  ws["A18"] = "Save Bonus"
    for i, (sc, mo, sp, sv) in enumerate(zip(
        c["scores"], c["mods"], c.get("save_prof", [None]*6), c["saves"]
    )):
        ws.cell(row=15, column=2+i, value=sc)
        ws.cell(row=16, column=2+i, value=mo)
        if sp: ws.cell(row=17, column=2+i, value="Y")
        ws.cell(row=18, column=2+i, value=sv)

    # Defenses
    ws["A20"] = "DEFENSES & SENSES"
    ws["A21"] = "Resistances";           ws["B21"] = c.get("resistances", "")
    ws["A22"] = "Immunities";            ws["B22"] = c.get("immunities", "")
    ws["A23"] = "Condition Immunities";  ws["B23"] = c.get("cond_immune", "")
    ws["A24"] = "Senses";                ws["B24"] = c.get("senses", "darkvision 60 ft., passive Perception 10")

    # Attacks
    ws["A26"] = "ATTACKS"
    for col, lbl in enumerate(["Name","Type","To Hit","Reach / Range","Damage","Save","Effect"], start=1):
        ws.cell(row=27, column=col, value=lbl)
    r = 28
    for nm, tp, hit, rng, dmg, sv, eff in c.get("attacks", []):
        ws.cell(row=r, column=1, value=nm)
        ws.cell(row=r, column=2, value=tp)
        ws.cell(row=r, column=3, value=hit)
        ws.cell(row=r, column=4, value=rng)
        ws.cell(row=r, column=5, value=dmg)
        ws.cell(row=r, column=6, value=sv)
        ws.cell(row=r, column=7, value=eff)
        r += 1
    next_row = r + 1

    for anchor, entries in [
        ("TRAITS",        c.get("traits",        [])),
        ("ACTIONS",       c.get("actions",       [])),
        ("BONUS ACTIONS", c.get("bonus_actions", [])),
        ("REACTIONS",     c.get("reactions",     [])),
        ("BLOODIED",      c.get("bloodied",      [])),
    ]:
        if entries:
            next_row = write_section(ws, next_row, anchor, entries)

    return ws


def add_battle_list_row(wb, tab, name=None, qty=1, group_name="",
                         init_mode="Individual", hp_override=None, notes=""):
    bl   = wb["Battle List"]
    name = name or tab
    for r in range(1, bl.max_row + 2):
        if bl.cell(row=r, column=1).value is None:
            bl.cell(row=r, column=1, value="NPC")
            bl.cell(row=r, column=2, value=name)
            bl.cell(row=r, column=3, value=tab)
            bl.cell(row=r, column=4, value=qty)
            bl.cell(row=r, column=5, value=group_name or None)
            bl.cell(row=r, column=6, value=init_mode)
            bl.cell(row=r, column=7, value=hp_override)
            bl.cell(row=r, column=8, value=notes or None)
            return r
    return None


# ─────────────────────────────────────────────────────────────────────────────
# NPC DEFINITIONS
# ─────────────────────────────────────────────────────────────────────────────

NPCS = [

    # ── 1. LINE SOLDIER ──────────────────────────────────────────────────────
    {
        "tab":          "Ganador Legionnaire",
        "epithet":      "Line Soldier",
        "role":         "Formation Fighter — holds line, controls space",
        "size_type":    "Medium humanoid (hobgoblin), lawful neutral",
        "ac": 16, "hp": 30, "speed": "30 ft.", "prof": 2, "init_mod": 1,
        "atk_per_round": 1,
        "scores":    [16, 12, 16, 10, 12, 10],
        "mods":      [ 3,  1,  3,  0,  1,  0],
        "save_prof": [False,False,True,False,True,False],
        "saves":     [ 3,  1,  5,  0,  3,  0],
        "resistances": "", "immunities": "",
        "cond_immune": "",
        "senses": "darkvision 60 ft., passive Perception 11",
        "attacks": [
            ("Spear Strike", "Melee Weapon", 5, "5 ft.", "1d8+3 piercing", "—", ""),
        ],
        "traits": [
            ("Formation Fighting",
             "+1 AC while adjacent to at least one ally (already included if in formation). "
             "Effective AC 17 in formation."),
            ("Disciplined",
             "Advantage on saving throws against being frightened."),
            ("Tactics",
             "Maintains formation; protects allies; does not pursue fleeing enemies. "
             "Break condition: if isolated, switches to full defense (no opportunity attacks, "
             "total defense action)."),
        ],
        "bonus_actions": [
            ("Brace",
             "Ready a reaction attack against the next creature that moves within 5 ft. "
             "of this soldier without using the Disengage action. On trigger, makes one "
             "Spear Strike as a reaction."),
        ],
        "bl_notes": "Backbone of Ganador — use in groups",
    },

    # ── 2. ARCHER ────────────────────────────────────────────────────────────
    {
        "tab":          "Ganador Marksman",
        "epithet":      "Archer",
        "role":         "Ranged Pressure — backline, avoids melee",
        "size_type":    "Medium humanoid (hobgoblin), lawful neutral",
        "ac": 14, "hp": 24, "speed": "30 ft.", "prof": 2, "init_mod": 3,
        "atk_per_round": 1,
        "scores":    [12, 16, 14, 10, 14, 10],
        "mods":      [ 1,  3,  2,  0,  2,  0],
        "save_prof": [False,True,False,False,True,False],
        "saves":     [ 1,  5,  2,  0,  4,  0],
        "resistances": "", "immunities": "", "cond_immune": "",
        "senses": "darkvision 60 ft., passive Perception 12",
        "attacks": [
            ("Longbow", "Ranged Weapon", 6, "150/600 ft.", "1d8+3 piercing", "—",
             "Volley Training: +2 to hit vs engaged targets (already factored in where applicable)"),
        ],
        "traits": [
            ("Volley Training",
             "+2 to hit against targets that are engaged in melee combat (i.e. within 5 ft. "
             "of an enemy). Effective +8 to hit in those cases."),
            ("Positioned",
             "+2 AC while in half or three-quarters cover."),
            ("Tactics",
             "Targets exposed or isolated enemies; repositions to maintain range. "
             "Break condition: disengages and withdraws if a melee threat closes to within 5 ft."),
        ],
        "actions": [
            ("Focused Shot",
             "Make one Longbow attack at disadvantage. On hit, deal an additional 2d6 piercing "
             "damage. (Attack +6, total damage 1d8+2d6+3 piercing.)"),
        ],
        "bl_notes": "Backline pressure — position on elevation when possible",
    },

    # ── 3. SERGEANT ──────────────────────────────────────────────────────────
    {
        "tab":          "Ganador Veteran",
        "epithet":      "Sergeant",
        "role":         "Unit Commander — boosts allies, holds center",
        "size_type":    "Medium humanoid (hobgoblin), lawful neutral",
        "ac": 17, "hp": 45, "speed": "30 ft.", "prof": 2, "init_mod": 1,
        "atk_per_round": 1,
        "scores":    [16, 12, 16, 12, 14, 14],
        "mods":      [ 3,  1,  3,  1,  2,  2],
        "save_prof": [False,False,True,False,True,False],
        "saves":     [ 3,  1,  5,  1,  4,  2],
        "resistances": "", "immunities": "", "cond_immune": "",
        "senses": "darkvision 60 ft., passive Perception 12",
        "attacks": [
            ("Heavy Strike", "Melee Weapon", 6, "5 ft.", "1d10+3 slashing", "—", ""),
        ],
        "traits": [
            ("Command Presence",
             "Friendly creatures within 10 ft. gain +1 to attack rolls while this sergeant "
             "is conscious and not incapacitated."),
            ("Veteran",
             "Once per round, can choose to gain advantage on one saving throw before rolling."),
            ("Tactics",
             "Stays in the center of the unit; reinforces weak points. "
             "Break condition: below half HP → retreats to preserve self and keep unit cohesion. "
             "Priority target for the party."),
        ],
        "bonus_actions": [
            ("Command",
             "One ally within 30 ft. that can hear the sergeant can use its reaction to "
             "move up to its speed or make one weapon attack."),
        ],
        "reactions": [
            ("Hold the Line",
             "When an ally within 5 ft. takes damage, the sergeant can use its reaction to "
             "reduce that damage by 1d10+3."),
        ],
        "bl_notes": "CRITICAL TARGET — killing the sergeant degrades unit performance",
    },

    # ── 4. WAR MAGE ──────────────────────────────────────────────────────────
    {
        "tab":          "War Mage",
        "epithet":      "Battlefield Caster",
        "role":         "Buff, control, disruption — stays protected",
        "size_type":    "Medium humanoid (hobgoblin or elf), lawful neutral",
        "ac": 13, "hp": 35, "speed": "30 ft.", "prof": 2, "init_mod": 2,
        "atk_per_round": 1,
        "scores":    [10, 14, 14, 18, 14, 12],
        "mods":      [ 0,  2,  2,  4,  2,  1],
        "save_prof": [False,False,True,True,False,False],
        "saves":     [ 0,  2,  4,  6,  2,  1],
        "resistances": "", "immunities": "", "cond_immune": "",
        "senses": "darkvision 60 ft., passive Perception 12",
        "attacks": [
            ("Fire Bolt",    "Ranged Spell", 7, "120 ft.", "2d10 fire",    "—",        "Cantrip — reliable ranged damage"),
            ("Ray of Frost", "Ranged Spell", 7, "60 ft.",  "1d8 cold",     "—",        "Cantrip — slows movement by 10 ft."),
            ("Magic Missile","Ranged Spell", "—","60 ft.", "3d4+3 force",  "—",        "1st level — three darts, auto-hit, guaranteed damage"),
            ("Scorching Ray","Ranged Spell", 7, "120 ft.", "2d6 fire",     "—",        "2nd level — three rays, each +7 to hit"),
            ("Fireball",     "Ranged Spell", "—","150 ft.","8d6 fire",     "DC 15 DEX","3rd level — 20-ft radius; half on save"),
            ("Lightning Bolt","Ranged Spell","—","100 ft.","8d6 lightning","DC 15 DEX","3rd level — 100-ft line, 5-ft wide; half on save"),
            ("Force Burst",  "Ranged Spell", 7, "60 ft.", "4d6 force",    "DC 15 CON","Utility — area force damage"),
        ],
        "traits": [
            ("Controlled Casting",
             "Not disrupted by damage less than 10 HP in a single hit (advantage on concentration saves)."),
            ("Battlefield Focus",
             "Advantage on Constitution saving throws to maintain concentration."),
            ("Spellcasting (Wizard ~7-9)",
             "Spell attack +7. Save DC 15. Spellcasting ability: Intelligence. "
             "Slots: 1st ×4, 2nd ×3, 3rd ×3, 4th ×1-2. "
             "Cantrips (at will): Fire Bolt, Ray of Frost, Shocking Grasp, Minor Illusion. "
             "1st: Mage Armor (pre-cast, gives AC 16), Shield (reaction +5 AC), Magic Missile, Fog Cloud. "
             "2nd: Mirror Image, Misty Step, Web, Scorching Ray. "
             "3rd: Fireball, Lightning Bolt, Haste (targets Shock Trooper or Heavy Infantry). "
             "4th: Wall of Fire (zone control), Greater Invisibility (protects key unit)."),
            ("Tactics",
             "Stays behind front line; uses Haste on high-value melee allies. Casts Fireball/Lightning Bolt "
             "into clusters. Pre-casts Mage Armor (replaces AC 13 with 16). "
             "Break condition: retreats immediately if threatened in melee; uses Misty Step to escape."),
        ],
        "actions": [
            ("Control Field",
             "Cast Web (DC 15 STR, restrained) or Slow (DC 15 WIS, -2 AC/-2 DEX saves, "
             "halved speed, one action or bonus action) in a 30–40 ft. area."),
            ("Buff Ally",
             "Cast Haste on one ally (doubled speed, +2 AC, advantage on DEX saves, extra Attack action). "
             "Concentration up to 1 minute."),
        ],
        "reactions": [
            ("Shield",
             "When hit by an attack or targeted by Magic Missile: +5 AC until start of next turn "
             "(including against the triggering attack). Uses a 1st-level slot."),
        ],
        "bl_notes": "HIGH VALUE TARGET — kill before it casts Haste or Fireball",
    },

    # ── 5. ARCANE SPECIALIST ─────────────────────────────────────────────────
    {
        "tab":          "Arcane Specialist",
        "epithet":      "Counter-Mage",
        "role":         "Anti-magic operative — targets enemy casters first",
        "size_type":    "Medium humanoid (hobgoblin or elf), lawful neutral",
        "ac": 14, "hp": 30, "speed": "30 ft.", "prof": 2, "init_mod": 2,
        "atk_per_round": 1,
        "scores":    [10, 14, 14, 18, 14, 12],
        "mods":      [ 0,  2,  2,  4,  2,  1],
        "save_prof": [False,False,False,True,True,False],
        "saves":     [ 0,  2,  2,  6,  4,  1],
        "resistances": "", "immunities": "", "cond_immune": "",
        "senses": "darkvision 60 ft., passive Perception 12",
        "attacks": [
            ("Arcane Strike",  "Melee/Ranged Spell", 5, "30 ft.", "1d6+2 force", "—",
             "On hit: target makes DC 14 CON save or loses concentration; "
             "if caster, next spell of 3rd level or lower fails (Counterspell-lite effect)."),
            ("Fire Bolt",      "Ranged Spell", 6, "120 ft.", "1d10 fire",    "—", "Cantrip"),
            ("Magic Missile",  "Ranged Spell", "—","60 ft.", "3d4+3 force",  "—", "1st level — auto-hit forces concentration check"),
            ("Hold Person",    "Ranged Spell", "—","60 ft.", "—",            "DC 14 WIS",
             "2nd level — target is paralyzed; attacks within 5 ft. are critical hits. Concentration."),
            ("Counterspell",   "Reaction",     "—","60 ft.", "—",            "—",
             "PRIMARY FUNCTION. Reaction when enemy casts a spell within 60 ft. "
             "Automatically negates spells 3rd level or lower. "
             "Higher: INT check DC = 10 + spell level. Uses 3rd-level slot."),
            ("Dispel Magic",   "Ranged Spell", "—","120 ft.","—",            "—",
             "3rd level — ends ongoing magical effects. DC = 10 + spell level for higher effects."),
            ("Lightning Bolt", "Ranged Spell", "—","100 ft.","8d6 lightning","DC 14 DEX",
             "3rd level — rare use; 100-ft line. Half on save."),
        ],
        "traits": [
            ("Detect Magic (Passive)",
             "Automatically senses the presence of magic within 30 ft. Knows schools and "
             "approximate locations. Cannot be surprised by magical effects."),
            ("Spellcasting (Wizard ~6-8)",
             "Spell attack +6. Save DC 14. Spellcasting ability: Intelligence. "
             "Slots: 1st ×4, 2nd ×3, 3rd ×3. "
             "Cantrips: Fire Bolt, Ray of Frost, Mage Hand, Minor Illusion. "
             "1st: Shield (reaction), Magic Missile, Detect Magic, Absorb Elements. "
             "2nd: Misty Step, Mirror Image, See Invisibility, Hold Person. "
             "3rd: Counterspell (primary!), Dispel Magic, Lightning Bolt (rare)."),
            ("Tactics",
             "Identifies and targets enemy spellcasters first. Hoards Counterspell slots "
             "for high-level spells. Uses See Invisibility vs stealthed casters. "
             "Break condition: if no magic threat is present, low-priority target "
             "— will withdraw rather than waste slots on mundane combat."),
        ],
        "actions": [
            ("Disrupt Casting",
             "Force one spellcaster within 30 ft. to make a DC 14 CON concentration check "
             "or lose their spell. This is not Counterspell — it happens before or during casting."),
        ],
        "reactions": [
            ("Counterspell",
             "When a creature within 60 ft. casts a spell: negate automatically if 3rd level "
             "or lower. For 4th level+: make DC 10 + spell_level INT check. Uses 3rd-level slot."),
            ("Shield",
             "+5 AC until start of next turn. Uses 1st-level slot."),
        ],
        "bl_notes": "Anti-PC specialist — priority target if party has casters",
    },

    # ── 6. WAR PRIEST ────────────────────────────────────────────────────────
    {
        "tab":          "War Priest",
        "epithet":      "Medic",
        "role":         "Battlefield sustain — keeps soldiers alive",
        "size_type":    "Medium humanoid (hobgoblin, human, or lizardfolk), lawful neutral",
        "ac": 15, "hp": 38, "speed": "30 ft.", "prof": 2, "init_mod": 0,
        "atk_per_round": 1,
        "scores":    [12, 10, 16, 10, 18, 14],
        "mods":      [ 1,  0,  3,  0,  4,  2],
        "save_prof": [False,False,True,False,True,False],
        "saves":     [ 1,  0,  5,  0,  6,  2],
        "resistances": "", "immunities": "", "cond_immune": "",
        "senses": "darkvision 60 ft., passive Perception 14",
        "attacks": [
            ("Light Strike",  "Melee Weapon", 5, "5 ft.", "1d6+2 bludgeoning", "—", "Mace or staff"),
            ("Sacred Flame",  "Ranged Spell", "—","60 ft.","1d8 radiant",       "DC 14 DEX",
             "Cantrip — no attack roll; target makes DEX save or takes damage. Ignores cover."),
        ],
        "traits": [
            ("Efficient Healing",
             "Healing spells restore the maximum die value when cast on a creature at or "
             "below half HP (DM option — simplifies tracking at the table)."),
            ("Stabilize",
             "As a bonus action, can touch a creature at 0 HP to stabilize it (no roll needed)."),
            ("Spellcasting (Cleric ~5-7)",
             "Spell attack +6. Save DC 14. Spellcasting ability: Wisdom. "
             "Slots: 1st ×4, 2nd ×3, 3rd ×2. "
             "Cantrips: Sacred Flame, Spare the Dying, Thaumaturgy, Guidance. "
             "1st: Cure Wounds (1d8+4), Healing Word (bonus action: 1d4+4), "
             "Bless (concentration, +1d4 to attacks and saves for 3 allies), Shield of Faith (+2 AC). "
             "2nd: Prayer of Healing (heal 2d8+4 to 6 targets — only outside combat), "
             "Spiritual Weapon (bonus action, 1d8+4 force, lasts 1 min). "
             "3rd: Mass Healing Word (bonus action, 1d4+4 to 6 targets), "
             "Revivify (restore 0 HP creature to 1 HP within 1 minute of death)."),
            ("Tactics",
             "Stays behind the front line; prioritizes wounded below half HP. "
             "Uses Healing Word (bonus action) to combine with a cantrip attack. "
             "Break condition: withdraws if threatened — not a frontline fighter."),
        ],
        "actions": [
            ("Heal Ally",
             "Cast Cure Wounds on a touched ally: restore 1d8+4 HP. "
             "Or Healing Word (bonus action, 30 ft.): restore 1d4+4 HP."),
            ("Bless",
             "Concentration, 1 min. Up to 3 creatures within 30 ft. add 1d4 to "
             "attack rolls and saving throws."),
        ],
        "bl_notes": "Sustain target — kill to prevent healing; priority after sergeant",
    },

    # ── 7. ILLUSIONIST ───────────────────────────────────────────────────────
    {
        "tab":          "Illusionist",
        "epithet":      "Field Deception Specialist",
        "role":         "Pre-battle misdirection, decoys, false terrain",
        "size_type":    "Medium humanoid (elf or hobgoblin), neutral",
        "ac": 13, "hp": 32, "speed": "30 ft.", "prof": 2, "init_mod": 2,
        "atk_per_round": 1,
        "scores":    [10, 14, 14, 18, 14, 12],
        "mods":      [ 0,  2,  2,  4,  2,  1],
        "save_prof": [False,False,False,True,True,False],
        "saves":     [ 0,  2,  2,  6,  4,  1],
        "resistances": "", "immunities": "", "cond_immune": "",
        "senses": "darkvision 60 ft., passive Perception 12",
        "attacks": [
            ("Fire Bolt",         "Ranged Spell", 6, "120 ft.", "1d10 fire",  "—",
             "Cantrip — fallback only; this specialist avoids combat"),
            ("Hypnotic Pattern",  "Ranged Spell", "—","120 ft.","—",          "DC 15 WIS",
             "3rd level — emergency crowd control. 30-ft cube. Creatures that fail are incapacitated "
             "and have speed 0. Concentration, 1 min. Breaks if target takes damage or is shaken."),
            ("Phantasmal Force",  "Ranged Spell", "—","60 ft.", "1d6 psychic","DC 15 INT",
             "2nd level — target believes an illusion is real; takes 1d6 psychic/round. Concentration."),
        ],
        "traits": [
            ("Illusionist School",
             "Can make one illusion spell appear completely real (Malleable Illusions). "
             "Minor Illusion can affect sound AND image simultaneously."),
            ("Nondetection (Passive)",
             "Protected from divination magic (Detect Magic, scrying, etc.) at all times. "
             "Can extend to one object or person as a 3rd-level spell."),
            ("Spellcasting (Wizard Illusion ~11)",
             "Spell attack +6. Save DC 15. Spellcasting ability: Intelligence. "
             "Slots: 1st ×4, 2nd ×3, 3rd ×3, 4th ×1, 5th ×1, 6th ×1. "
             "Cantrips: Minor Illusion (sound+image), Mage Hand, Message, Fire Bolt. "
             "1st: Disguise Self, Silent Image (15-ft cube, concentration), Fog Cloud. "
             "2nd: Mirror Image, Invisibility, Phantasmal Force. "
             "3rd: Major Image (20-ft cube, permanent once/day — primary battlefield deception), "
             "Hypnotic Pattern, Nondetection. "
             "4th: Greater Invisibility, Hallucinatory Terrain (pre-battle setup). "
             "5th/6th: reserved for Mislead or Programmed Illusion."),
            ("Tactics",
             "Pre-battle: Hallucinatory Terrain to reshape apparent environment. "
             "During combat: Major Image for decoys and false flanks (permanent after 10 rounds, once/day). "
             "Invisibility on self or key ally to reposition. Hypnotic Pattern only as emergency. "
             "Indistinguishable from other personnel — may not be identified as a caster until casting. "
             "Break condition: immediately retreats using Invisibility or Misty Step if discovered."),
        ],
        "actions": [
            ("Major Image",
             "Creates an illusory object, creature, or phenomenon up to a 20-ft cube within 120 ft. "
             "Affects sight, sound, smell, and temperature. Concentration, 10 minutes. "
             "Once per day: if maintained for the full duration, becomes permanent (no concentration)."),
            ("Disguise Self",
             "Changes apparent appearance, clothing, and equipment. Lasts 1 hour. "
             "Physical interaction or DC 15 Investigation check reveals the illusion."),
        ],
        "bl_notes": "May be unidentified as a caster — watch for out-of-place NPCs",
    },

    # ── 8. FIRE ANCHOR CASTER ────────────────────────────────────────────────
    {
        "tab":          "Fire Anchor Caster",
        "epithet":      "Elemental Ritual Cultist",
        "role":         "Maintains fire portal connection — does not fight",
        "size_type":    "Medium humanoid (cultist), chaotic neutral",
        "ac": 12, "hp": 18, "speed": "30 ft.", "prof": 2, "init_mod": 0,
        "atk_per_round": 0,
        "scores":    [10, 10, 12, 12, 14, 14],
        "mods":      [ 0,  0,  1,  1,  2,  2],
        "save_prof": [False,False,False,False,True,True],
        "saves":     [ 0,  0,  1,  1,  4,  4],
        "resistances": "fire (while portal Stage 2+)",
        "immunities": "", "cond_immune": "",
        "senses": "passive Perception 10",
        "attacks": [],
        "traits": [
            ("Elemental Ward",
             "While the portal is Stage 2 or higher: advantage on saves against elemental effects, "
             "and resistance to all elemental damage types."),
            ("Ritual Trance",
             "Does not make attacks, cast spells, or take opportunity attacks. "
             "Disadvantage on Perception checks while in trance. Cannot be reasoned with while chanting."),
            ("Bound to Fire Focus",
             "The fire portal connection persists until this caster is stopped (killed/incapacitated/"
             "removed from the circle) AND the fire focus is stolen, disabled, dispelled, or destroyed."),
            ("Nonviolent Disruption",
             "The caster can be stopped without killing: "
             "DC 16 Persuasion, Intimidation, Religion, or Arcana (break the chant); "
             "DC 15 Athletics (drag from the circle); "
             "Dispel Magic vs DC 16; "
             "Silence, stun, paralysis, or incapacitation also works."),
            ("Elemental Focus",
             "Small magical object. AC 15, HP 20, Damage Threshold 5. "
             "Immune to poison and psychic. Resistant to elemental damage (Stage 2+). "
             "Can be stolen (Sleight of Hand DC 16 or Athletics DC 15), "
             "disabled (Arcana/Religion DC 16), or dispelled (DC 16). "
             "Covered or contained: DM call, DC 15-16 appropriate action."),
        ],
        "actions": [
            ("Continue Ritual",
             "Maintains the red/fire connection to the elemental portal. "
             "No other action available while in Ritual Trance."),
            ("Cry Out",
             "Trigger: when this caster is attacked or the focus is touched. "
             "One defender or elemental that can hear this caster may immediately move "
             "10 ft. toward this anchor as a reaction."),
        ],
        "bloodied": [
            ("Violent Backlash — if killed or focus destroyed",
             "Roll d6 when the anchor connection ends violently. "
             "1–2: Fire Detonation — 20-ft radius, DC 16 DEX save, 8d6 fire damage. "
             "3–4: Fire Birth — a fire elemental appears at this anchor location. "
             "5–6: Fire Rift — a fire fountain persists; creatures entering or starting "
             "their turn within 10 ft. take 2d6 fire damage."),
        ],
        "bl_notes": "Ritual target — focus on disruption, not damage",
    },

    # ── 9. EARTH ANCHOR CASTER ───────────────────────────────────────────────
    {
        "tab":          "Earth Anchor Caster",
        "epithet":      "Elemental Ritual Cultist",
        "role":         "Maintains earth portal connection — does not fight",
        "size_type":    "Medium humanoid (cultist), chaotic neutral",
        "ac": 12, "hp": 18, "speed": "30 ft.", "prof": 2, "init_mod": 0,
        "atk_per_round": 0,
        "scores":    [10, 10, 12, 12, 14, 14],
        "mods":      [ 0,  0,  1,  1,  2,  2],
        "save_prof": [False,False,False,False,True,True],
        "saves":     [ 0,  0,  1,  1,  4,  4],
        "resistances": "earth/bludgeoning (while portal Stage 2+)",
        "immunities": "", "cond_immune": "",
        "senses": "passive Perception 10",
        "attacks": [],
        "traits": [
            ("Elemental Ward",
             "While the portal is Stage 2 or higher: advantage on saves against elemental effects, "
             "and resistance to all elemental damage types."),
            ("Ritual Trance",
             "Does not make attacks, cast spells, or take opportunity attacks. "
             "Disadvantage on Perception checks while in trance."),
            ("Bound to Earth Focus",
             "The earth portal connection persists until this caster is stopped AND the earth "
             "focus is stolen, disabled, dispelled, or destroyed."),
            ("Nonviolent Disruption",
             "DC 16 Persuasion, Intimidation, Religion, or Arcana (break the chant); "
             "DC 15 Athletics (drag from the circle); Dispel Magic vs DC 16; "
             "Silence, stun, paralysis, or incapacitation also works."),
            ("Elemental Focus",
             "AC 15, HP 20, Damage Threshold 5. Can be stolen (SoH DC 16 or Athletics DC 15), "
             "disabled (Arcana/Religion DC 16), covered (DC 15-16), or dispelled (DC 16)."),
        ],
        "actions": [
            ("Continue Ritual",
             "Maintains the orange/earth connection to the elemental portal."),
            ("Cry Out",
             "Trigger: when attacked or focus touched. One defender or elemental "
             "within hearing may move 10 ft. toward this anchor as a reaction."),
        ],
        "bloodied": [
            ("Violent Backlash — if killed or focus destroyed",
             "Roll d6: "
             "1–2: Stone Detonation — 20-ft radius, DC 16 DEX or STR save, 8d6 bludgeoning; prone on fail. "
             "3–4: Stone Birth — an earth elemental appears at this anchor location. "
             "5–6: Earth Rift — cracked ground remains as difficult terrain; "
             "DC 16 DEX or fall into a 10-ft crack."),
        ],
        "bl_notes": "Ritual target — orange connection; focus on disruption, not damage",
    },

    # ── 10. WATER ANCHOR CASTER ──────────────────────────────────────────────
    {
        "tab":          "Water Anchor Caster",
        "epithet":      "Elemental Ritual Cultist",
        "role":         "Maintains water portal connection — does not fight",
        "size_type":    "Medium humanoid (cultist), chaotic neutral",
        "ac": 12, "hp": 18, "speed": "30 ft.", "prof": 2, "init_mod": 0,
        "atk_per_round": 0,
        "scores":    [10, 10, 12, 12, 14, 14],
        "mods":      [ 0,  0,  1,  1,  2,  2],
        "save_prof": [False,False,False,False,True,True],
        "saves":     [ 0,  0,  1,  1,  4,  4],
        "resistances": "cold, water (while portal Stage 2+)",
        "immunities": "", "cond_immune": "",
        "senses": "passive Perception 10",
        "attacks": [],
        "traits": [
            ("Elemental Ward",
             "While the portal is Stage 2 or higher: advantage on elemental saves, "
             "resistance to all elemental damage types."),
            ("Ritual Trance",
             "Does not attack, cast spells, or take opportunity attacks. "
             "Disadvantage on Perception."),
            ("Bound to Water Focus",
             "The water portal connection persists until this caster is stopped AND the water "
             "focus is stolen, disabled, dispelled, or destroyed."),
            ("Nonviolent Disruption",
             "DC 16 Persuasion, Intimidation, Religion, or Arcana; DC 15 Athletics; "
             "Dispel Magic DC 16; Silence/stun/paralysis/incapacitation."),
            ("Elemental Focus", "AC 15, HP 20, Damage Threshold 5."),
        ],
        "actions": [
            ("Continue Ritual",
             "Maintains the blue/water connection to the elemental portal."),
            ("Cry Out",
             "One defender or elemental within hearing moves 10 ft. toward this anchor as a reaction."),
        ],
        "bloodied": [
            ("Violent Backlash — if killed or focus destroyed",
             "Roll d6: "
             "1–2: Steam Detonation — 20-ft radius, DC 16 CON save, 8d6 scalding damage; "
             "blinded until end of next turn on fail. "
             "3–4: Water Birth — a water elemental appears. "
             "5–6: Water Rift — flood/wall remains as difficult terrain; "
             "DC 16 STR or pushed 10 ft."),
        ],
        "bl_notes": "Ritual target — blue connection",
    },

    # ── 11. AIR ANCHOR CASTER ────────────────────────────────────────────────
    {
        "tab":          "Air Anchor Caster",
        "epithet":      "Elemental Ritual Cultist",
        "role":         "Maintains air portal connection — does not fight",
        "size_type":    "Medium humanoid (cultist), chaotic neutral",
        "ac": 12, "hp": 18, "speed": "30 ft.", "prof": 2, "init_mod": 0,
        "atk_per_round": 0,
        "scores":    [10, 10, 12, 12, 14, 14],
        "mods":      [ 0,  0,  1,  1,  2,  2],
        "save_prof": [False,False,False,False,True,True],
        "saves":     [ 0,  0,  1,  1,  4,  4],
        "resistances": "lightning, thunder (while portal Stage 2+)",
        "immunities": "", "cond_immune": "",
        "senses": "passive Perception 10",
        "attacks": [],
        "traits": [
            ("Elemental Ward",
             "While the portal is Stage 2 or higher: advantage on elemental saves, "
             "resistance to all elemental damage types."),
            ("Ritual Trance",
             "Does not attack, cast spells, or take opportunity attacks. "
             "Disadvantage on Perception."),
            ("Bound to Air Focus",
             "The air portal connection persists until this caster is stopped AND the air "
             "focus is stolen, disabled, dispelled, or destroyed."),
            ("Nonviolent Disruption",
             "DC 16 Persuasion, Intimidation, Religion, or Arcana; DC 15 Athletics; "
             "Dispel Magic DC 16; Silence/stun/paralysis/incapacitation."),
            ("Elemental Focus", "AC 15, HP 20, Damage Threshold 5."),
        ],
        "actions": [
            ("Continue Ritual",
             "Maintains the yellow/air connection to the elemental portal."),
            ("Cry Out",
             "One defender or elemental within hearing moves 10 ft. toward this anchor as a reaction."),
        ],
        "bloodied": [
            ("Violent Backlash — if killed or focus destroyed",
             "Roll d6: "
             "1–2: Wind Detonation — 20-ft radius, DC 16 STR save, 8d6 force damage; "
             "pushed 20 ft. and knocked prone on fail. "
             "3–4: Air Birth — an air elemental appears. "
             "5–6: Air Rift — wind wall remains; ranged attacks through it have disadvantage; "
             "DC 16 STR to cross or be pushed 15 ft."),
        ],
        "bl_notes": "Ritual target — yellow connection",
    },
]

# ─────────────────────────────────────────────────────────────────────────────
# BATTLE LIST DEFAULTS
# ─────────────────────────────────────────────────────────────────────────────
BATTLE_LIST_ENTRIES = [
    # (tab_name, qty, init_mode, notes)
    ("Ganador Legionnaire",  1, "Individual", "Line soldier — add qty as needed"),
    ("Ganador Marksman",     1, "Individual", "Archer — position on elevation"),
    ("Ganador Veteran",      1, "Individual", "Sergeant — PRIORITY TARGET"),
    ("War Mage",             1, "Individual", "PRIORITY TARGET — kill before Haste/Fireball"),
    ("Arcane Specialist",    1, "Individual", "Anti-PC caster — watch for Counterspell"),
    ("War Priest",           1, "Individual", "Medic — kill to stop healing"),
    ("Illusionist",          1, "Individual", "May be disguised — watch for illusion signs"),
    ("Fire Anchor Caster",   1, "Individual", "Ritual target — does not fight"),
    ("Earth Anchor Caster",  1, "Individual", "Ritual target — does not fight"),
    ("Water Anchor Caster",  1, "Individual", "Ritual target — does not fight"),
    ("Air Anchor Caster",    1, "Individual", "Ritual target — does not fight"),
]

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if not os.path.exists(WORKBOOK):
        print(f"ERROR: Workbook not found: {WORKBOOK}")
        sys.exit(1)

    # Safe write: load → modify /tmp copy → replace original
    print(f"Loading {WORKBOOK} ...")
    shutil.copy2(WORKBOOK, TMP)
    wb = openpyxl.load_workbook(TMP)

    for npc in NPCS:
        add_creature_tab(wb, npc)
        print(f"  Added tab: {npc['tab']}")

    for tab, qty, init_mode, notes in BATTLE_LIST_ENTRIES:
        row = add_battle_list_row(wb, tab, qty=qty, init_mode=init_mode, notes=notes)
        print(f"  Battle List row {row}: {tab}")

    wb.save(TMP)
    shutil.copy2(TMP, WORKBOOK)
    print(f"\nDone. {len(NPCS)} NPC tabs added to {WORKBOOK}")
    print("Sheets:", [s for s in openpyxl.load_workbook(WORKBOOK).sheetnames])
