"""
add_monsters.py — Add a "Whitman sampler" bestiary of SRD monsters to
                  combat tracker.xlsx as creature tabs + Battle List rows.

Adds: elementals (fire/earth/air/water), mephits (incl. magma), undead
(skeleton/zombie/shadow/ghoul/wight/specter/wraith), humanoid raiders
(goblin/orc/ogre/bugbear/orc war chief), giants (fire/storm/frost/stone),
and the five adult chromatic dragons (red/blue/green/black/white).

Format matches engine/excel_loader.py EXACTLY (fixed-cell stat blocks +
section anchors). Helper functions are lifted verbatim from add_npcs.py.

SAFE WRITE: copies the workbook to /tmp, modifies the copy, then replaces the
original. Existing sheets are NEVER deleted unless a new tab has the same name
(none of these collide with existing tabs). Battle List rows are appended with
Quantity 0 (library entries — bump to >0 to drop one into a fight), matching
the existing Ganador/anchor convention.

RUN from the battle tracker folder:  python3 add_monsters.py
"""
import openpyxl
import shutil
import os
import sys

# Resolve workbook path relative to this script so it works on any machine.
HERE = os.path.dirname(os.path.abspath(__file__))
WORKBOOK = os.path.join(HERE, "combat tracker.xlsx")
TMP      = "/tmp/combat_tracker_monsters.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS (verbatim format from add_npcs.py)
# ─────────────────────────────────────────────────────────────────────────────

def write_section(ws, start_row, anchor, entries):
    ws.cell(row=start_row, column=1, value=anchor)
    r = start_row + 1
    for name, desc in entries:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=desc)
        r += 1
    return r + 1   # blank separator row


def add_creature_tab(wb, c):
    tab = c["tab"]
    if tab in wb.sheetnames:
        del wb[tab]
    ws = wb.create_sheet(tab)

    ws["A1"] = "CREATURE STAT BLOCK"
    ws["K2"] = c.get("atk_per_round", 1)

    ws["A3"]  = "Name";                    ws["B3"]  = tab
    ws["A4"]  = "Epithet";                 ws["B4"]  = c.get("epithet", "")
    ws["A5"]  = "Role";                    ws["B5"]  = c.get("role", "")
    ws["A6"]  = "Size / Type / Alignment"; ws["B6"]  = c.get("size_type", "")
    ws["A7"]  = "AC";                      ws["B7"]  = c["ac"]
    ws["A8"]  = "Base HP";                 ws["B8"]  = c["hp"]
    ws["A9"]  = "Speed";                   ws["B9"]  = c.get("speed", "30 ft.")
    ws["A10"] = "Proficiency Bonus";       ws["B10"] = c.get("prof", 2)
    ws["A11"] = "Initiative Mod";          ws["B11"] = c.get("init_mod", 0)

    # Ability scores
    ws["A13"] = "ABILITY SCORES"
    for i, lbl in enumerate(["STR", "DEX", "CON", "INT", "WIS", "CHA"]):
        ws.cell(row=14, column=2 + i, value=lbl)
    ws["A15"] = "Score";            ws["A16"] = "Modifier"
    ws["A17"] = "Save Proficient";  ws["A18"] = "Save Bonus"
    for i, (sc, mo, sp, sv) in enumerate(zip(
        c["scores"], c["mods"], c.get("save_prof", [False] * 6), c["saves"]
    )):
        ws.cell(row=15, column=2 + i, value=sc)
        ws.cell(row=16, column=2 + i, value=mo)
        if sp:
            ws.cell(row=17, column=2 + i, value="Y")
        ws.cell(row=18, column=2 + i, value=sv)

    # Defenses
    ws["A20"] = "DEFENSES & SENSES"
    ws["A21"] = "Resistances";           ws["B21"] = c.get("resistances", "")
    ws["A22"] = "Immunities";            ws["B22"] = c.get("immunities", "")
    ws["A23"] = "Condition Immunities";  ws["B23"] = c.get("cond_immune", "")
    ws["A24"] = "Senses";                ws["B24"] = c.get("senses", "passive Perception 10")

    # Attacks
    ws["A26"] = "ATTACKS"
    for col, lbl in enumerate(["Name", "Type", "To Hit", "Reach / Range", "Damage", "Save", "Effect"], start=1):
        ws.cell(row=27, column=col, value=lbl)
    r = 28
    for nm, tp, hit, rng, dmg, sv, eff in c.get("attacks", []):
        ws.cell(row=r, column=1, value=nm)
        ws.cell(row=r, column=2, value=tp)
        ws.cell(row=r, column=3, value=hit)
        ws.cell(row=r, column=4, value=rng)
        ws.cell(row=r, column=5, value=dmg)
        ws.cell(row=r, column=6, value=sv)
        ws.cell(row=r, column=7, value=eff)
        r += 1
    next_row = r + 1

    for anchor, entries in [
        ("TRAITS",        c.get("traits",        [])),
        ("ACTIONS",       c.get("actions",       [])),
        ("BONUS ACTIONS", c.get("bonus_actions", [])),
        ("REACTIONS",     c.get("reactions",     [])),
        ("BLOODIED",      c.get("bloodied",      [])),
    ]:
        if entries:
            next_row = write_section(ws, next_row, anchor, entries)

    return ws


def add_battle_list_row(wb, tab, name=None, qty=0, group_name="",
                        init_mode="Individual", hp_override=None, notes=""):
    bl = wb["Battle List"]
    name = name or tab
    # Skip if a row already references this tab (idempotent re-runs).
    for r in range(1, bl.max_row + 1):
        if bl.cell(row=r, column=3).value == tab:
            return None
    for r in range(1, bl.max_row + 2):
        if bl.cell(row=r, column=1).value is None:
            bl.cell(row=r, column=1, value="NPC")
            bl.cell(row=r, column=2, value=name)
            bl.cell(row=r, column=3, value=tab)
            bl.cell(row=r, column=4, value=qty)
            bl.cell(row=r, column=5, value=group_name or None)
            bl.cell(row=r, column=6, value=init_mode)
            bl.cell(row=r, column=7, value=hp_override)
            bl.cell(row=r, column=8, value=notes or None)
            return r
    return None


# ─────────────────────────────────────────────────────────────────────────────
# MONSTER DEFINITIONS  (stats from the 5e SRD / open content)
# ─────────────────────────────────────────────────────────────────────────────

CI_ELEMENTAL = "exhaustion, grappled, paralyzed, petrified, poisoned, prone, restrained, unconscious"
NONMAG_BPS   = "bludgeoning, piercing, slashing from nonmagical attacks"

MONSTERS = [

    # ════════════════════ ELEMENTALS ════════════════════
    {
        "tab": "Fire Elemental", "epithet": "Living Conflagration",
        "role": "Aggressive bruiser — ignites everything it touches",
        "size_type": "Large elemental, neutral",
        "ac": 13, "hp": 102, "speed": "50 ft.", "prof": 3, "init_mod": 3, "atk_per_round": 2,
        "scores": [10, 17, 16, 6, 10, 7], "mods": [0, 3, 3, -2, 0, -2],
        "save_prof": [False]*6, "saves": [0, 3, 3, -2, 0, -2],
        "resistances": NONMAG_BPS, "immunities": "fire, poison", "cond_immune": CI_ELEMENTAL,
        "senses": "darkvision 60 ft., passive Perception 10",
        "attacks": [
            ("Touch", "Melee Weapon", 6, "5 ft.", "2d6+3 fire", "—",
             "If target is flammable it ignites: 1d10 fire at the start of each of its turns until doused."),
        ],
        "traits": [
            ("Multiattack", "The elemental makes two Touch attacks."),
            ("Fire Form", "Can move through a 1-inch space without squeezing. A creature that touches it or "
             "hits it with a melee attack within 5 ft. takes 1d10 fire. It ignites flammable objects it touches."),
            ("Illumination", "Sheds bright light in a 30-ft radius and dim light for an extra 30 ft."),
            ("Water Susceptibility", "Takes 1 cold damage for every 5 ft. it moves in water, or per gallon splashed on it."),
        ],
        "bl_notes": "Elemental — fire portal / summon",
    },
    {
        "tab": "Water Elemental", "epithet": "Surging Tide",
        "role": "Grappler — engulfs and drowns",
        "size_type": "Large elemental, neutral",
        "ac": 14, "hp": 114, "speed": "30 ft., swim 90 ft.", "prof": 3, "init_mod": 2, "atk_per_round": 2,
        "scores": [18, 14, 18, 5, 10, 8], "mods": [4, 2, 4, -3, 0, -1],
        "save_prof": [False]*6, "saves": [4, 2, 4, -3, 0, -1],
        "resistances": "acid; " + NONMAG_BPS, "immunities": "poison", "cond_immune": CI_ELEMENTAL,
        "senses": "darkvision 60 ft., passive Perception 10",
        "attacks": [
            ("Slam", "Melee Weapon", 7, "5 ft.", "2d8+4 bludgeoning", "—", ""),
            ("Whelm", "Special (Recharge 4-6)", "—", "5 ft.", "2d8+4 bludgeoning", "DC 15 STR",
             "Engulfs Large-or-smaller creatures in its space (half damage on save). Engulfed: restrained, can't "
             "breathe, takes 2d8+4 bludgeoning at the start of each of the elemental's turns."),
        ],
        "traits": [
            ("Multiattack", "The elemental makes two Slam attacks."),
            ("Water Form", "Can move through a 1-inch space without squeezing."),
            ("Freeze", "If it takes cold damage, its speed is reduced by 20 ft. until the end of its next turn."),
        ],
        "bl_notes": "Elemental — water portal / summon",
    },
    {
        "tab": "Air Elemental", "epithet": "Howling Gale",
        "role": "Fast skirmisher — whirlwind through clusters",
        "size_type": "Large elemental, neutral",
        "ac": 15, "hp": 90, "speed": "0 ft., fly 90 ft. (hover)", "prof": 3, "init_mod": 5, "atk_per_round": 2,
        "scores": [14, 20, 14, 6, 10, 6], "mods": [2, 5, 2, -2, 0, -2],
        "save_prof": [False]*6, "saves": [2, 5, 2, -2, 0, -2],
        "resistances": "lightning, thunder; " + NONMAG_BPS, "immunities": "poison", "cond_immune": CI_ELEMENTAL,
        "senses": "darkvision 60 ft., passive Perception 10",
        "attacks": [
            ("Slam", "Melee Weapon", 8, "5 ft.", "2d8+5 bludgeoning", "—", ""),
            ("Whirlwind", "Special (Recharge 4-6)", "—", "self", "3d8+2 bludgeoning", "DC 13 STR",
             "Each creature in the elemental's space (half on save); on a failed save, target is flung up to 20 ft. "
             "away and knocked prone."),
        ],
        "traits": [
            ("Multiattack", "The elemental makes two Slam attacks."),
            ("Air Form", "Can move through a space as narrow as 1 inch without squeezing."),
        ],
        "bl_notes": "Elemental — air portal / summon",
    },
    {
        "tab": "Earth Elemental", "epithet": "Walking Mountain",
        "role": "Tank — slow, durable, hits like a landslide",
        "size_type": "Large elemental, neutral",
        "ac": 17, "hp": 126, "speed": "30 ft., burrow 30 ft.", "prof": 3, "init_mod": -1, "atk_per_round": 2,
        "scores": [20, 8, 20, 5, 10, 5], "mods": [5, -1, 5, -3, 0, -3],
        "save_prof": [False]*6, "saves": [5, -1, 5, -3, 0, -3],
        "resistances": NONMAG_BPS, "immunities": "poison", "cond_immune": CI_ELEMENTAL,
        "senses": "darkvision 60 ft., tremorsense 60 ft., passive Perception 10",
        "attacks": [
            ("Slam", "Melee Weapon", 8, "10 ft.", "2d8+5 bludgeoning", "—", ""),
        ],
        "traits": [
            ("Multiattack", "The elemental makes two Slam attacks."),
            ("Earth Glide", "Can burrow through nonmagical earth and stone without disturbing it."),
            ("Siege Monster", "Deals double damage to objects and structures."),
            ("Vulnerability", "Vulnerable to thunder damage."),
        ],
        "bl_notes": "Elemental — earth portal / summon",
    },

    # ════════════════════ MEPHITS ════════════════════
    {
        "tab": "Magma Mephit", "epithet": "Ember Imp",
        "role": "Minion skirmisher — explodes on death",
        "size_type": "Small elemental, neutral evil",
        "ac": 11, "hp": 22, "speed": "30 ft., fly 30 ft.", "prof": 2, "init_mod": 1, "atk_per_round": 1,
        "scores": [8, 12, 12, 7, 10, 10], "mods": [-1, 1, 1, -2, 0, 0],
        "save_prof": [False]*6, "saves": [-1, 1, 1, -2, 0, 0],
        "resistances": "", "immunities": "fire, poison", "cond_immune": "poisoned",
        "senses": "darkvision 60 ft., passive Perception 10", "vulnerable": "cold",
        "attacks": [
            ("Claws", "Melee Weapon", 3, "5 ft.", "1d4 slashing plus 1d4 fire", "—", ""),
            ("Fire Breath", "Special (Recharge 6)", "—", "15-ft cone", "2d6 fire", "DC 11 DEX", "Half on save."),
        ],
        "traits": [
            ("Death Burst", "When it dies, it explodes: each creature within 5 ft. makes a DC 11 DEX save, taking 1d8 fire on a fail."),
            ("False Appearance", "While motionless, indistinguishable from a lump of magma."),
            ("Vulnerability", "Vulnerable to cold damage."),
            ("Innate Spellcasting (1/day)", "Can cast Heat Metal (DC 10)."),
        ],
        "bl_notes": "Mephit — magma (CR 1/2)",
    },
    {
        "tab": "Steam Mephit", "epithet": "Scald Imp",
        "role": "Minion skirmisher — boiling cloud",
        "size_type": "Small elemental, neutral evil",
        "ac": 10, "hp": 21, "speed": "30 ft., fly 30 ft.", "prof": 2, "init_mod": 0, "atk_per_round": 1,
        "scores": [5, 11, 10, 11, 10, 12], "mods": [-3, 0, 0, 0, 0, 1],
        "save_prof": [False]*6, "saves": [-3, 0, 0, 0, 0, 1],
        "resistances": "", "immunities": "fire, poison", "cond_immune": "poisoned",
        "senses": "darkvision 60 ft., passive Perception 10",
        "attacks": [
            ("Claws", "Melee Weapon", 2, "5 ft.", "1d4 slashing plus 1d4 fire", "—", ""),
            ("Steam Breath", "Special (Recharge 6)", "—", "15-ft cone", "2d6 fire", "DC 10 DEX", "Half on save."),
        ],
        "traits": [
            ("Death Burst", "When it dies, it explodes in scalding steam: each creature within 5 ft. makes a DC 10 DEX save, taking 1d8 fire on a fail."),
            ("Innate Spellcasting (1/day)", "Can cast Blur (DC 11)."),
        ],
        "bl_notes": "Mephit — steam (CR 1/4)",
    },
    {
        "tab": "Ice Mephit", "epithet": "Frost Imp",
        "role": "Minion skirmisher — shatters into shards",
        "size_type": "Small elemental, neutral evil",
        "ac": 11, "hp": 21, "speed": "30 ft., fly 30 ft.", "prof": 2, "init_mod": 1, "atk_per_round": 1,
        "scores": [7, 13, 10, 9, 11, 12], "mods": [-2, 1, 0, -1, 0, 1],
        "save_prof": [False]*6, "saves": [-2, 1, 0, -1, 0, 1],
        "resistances": "", "immunities": "cold, poison", "cond_immune": "poisoned",
        "senses": "darkvision 60 ft., passive Perception 11", "vulnerable": "bludgeoning, fire",
        "attacks": [
            ("Claws", "Melee Weapon", 3, "5 ft.", "1d4 slashing plus 1d4 cold", "—", ""),
            ("Frost Breath", "Special (Recharge 6)", "—", "15-ft cone", "2d4 cold", "DC 10 DEX", "Half on save."),
        ],
        "traits": [
            ("Death Burst", "When it dies, it shatters: each creature within 5 ft. makes a DC 10 DEX save, taking 1d8 slashing on a fail."),
            ("False Appearance", "While motionless, indistinguishable from a lump of ice."),
            ("Vulnerability", "Vulnerable to bludgeoning and fire damage."),
            ("Innate Spellcasting (1/day)", "Can cast Fog Cloud."),
        ],
        "bl_notes": "Mephit — ice (CR 1/2)",
    },
    {
        "tab": "Dust Mephit", "epithet": "Grit Imp",
        "role": "Minion skirmisher — blinding cloud",
        "size_type": "Small elemental, neutral evil",
        "ac": 12, "hp": 17, "speed": "30 ft., fly 30 ft.", "prof": 2, "init_mod": 2, "atk_per_round": 1,
        "scores": [5, 14, 10, 9, 11, 10], "mods": [-3, 2, 0, -1, 0, 0],
        "save_prof": [False]*6, "saves": [-3, 2, 0, -1, 0, 0],
        "resistances": "", "immunities": "poison", "cond_immune": "poisoned",
        "senses": "darkvision 60 ft., passive Perception 10", "vulnerable": "fire",
        "attacks": [
            ("Claws", "Melee Weapon", 4, "5 ft.", "1d4+2 slashing", "—", ""),
            ("Blinding Breath", "Special (Recharge 6)", "—", "15-ft cone", "—", "DC 10 CON",
             "On a failed save, target is blinded until the end of the mephit's next turn."),
        ],
        "traits": [
            ("Death Burst", "When it dies, it bursts into a cloud of dust: each creature within 5 ft. makes a DC 10 CON save or is blinded until the end of its next turn."),
            ("Vulnerability", "Vulnerable to fire damage."),
            ("Innate Spellcasting (1/day)", "Can cast Sleep (DC 11)."),
        ],
        "bl_notes": "Mephit — dust (CR 1/2)",
    },
    {
        "tab": "Mud Mephit", "epithet": "Sludge Imp",
        "role": "Minion controller — restrains in muck",
        "size_type": "Small elemental, neutral evil",
        "ac": 11, "hp": 27, "speed": "20 ft., fly 20 ft., swim 20 ft.", "prof": 2, "init_mod": 0, "atk_per_round": 1,
        "scores": [8, 10, 12, 9, 11, 7], "mods": [-1, 0, 1, -1, 0, -2],
        "save_prof": [False]*6, "saves": [-1, 0, 1, -1, 0, -2],
        "resistances": "", "immunities": "poison", "cond_immune": "poisoned",
        "senses": "darkvision 60 ft., passive Perception 10",
        "attacks": [
            ("Fists", "Melee Weapon", 3, "5 ft.", "1d6+1 bludgeoning", "—", ""),
            ("Mud Breath", "Special (Recharge 6)", "—", "5-ft cone", "—", "DC 11 DEX",
             "On a failed save, target is restrained until the end of the mephit's next turn."),
        ],
        "traits": [
            ("Death Burst", "When it dies, it collapses into a pool of mud (no damage)."),
            ("False Appearance", "While motionless, indistinguishable from a pile of mud."),
        ],
        "bl_notes": "Mephit — mud (CR 1/4)",
    },
    {
        "tab": "Smoke Mephit", "epithet": "Cinder Imp",
        "role": "Minion skirmisher — choking cloud",
        "size_type": "Small elemental, neutral evil",
        "ac": 12, "hp": 22, "speed": "30 ft., fly 30 ft.", "prof": 2, "init_mod": 2, "atk_per_round": 1,
        "scores": [6, 14, 12, 10, 10, 11], "mods": [-2, 2, 1, 0, 0, 0],
        "save_prof": [False]*6, "saves": [-2, 2, 1, 0, 0, 0],
        "resistances": "", "immunities": "fire, poison", "cond_immune": "poisoned",
        "senses": "darkvision 60 ft., passive Perception 10",
        "attacks": [
            ("Claws", "Melee Weapon", 4, "5 ft.", "1d4+2 slashing", "—", ""),
            ("Cinder Breath", "Special (Recharge 6)", "—", "15-ft cone", "—", "DC 11 DEX",
             "On a failed save, target is blinded until the end of the mephit's next turn."),
        ],
        "traits": [
            ("Death Burst", "When it dies, it leaves a 5-ft cloud of smoke that lightly obscures its space."),
            ("Innate Spellcasting (1/day)", "Can cast Dancing Lights."),
        ],
        "bl_notes": "Mephit — smoke (CR 1/4)",
    },

    # ════════════════════ UNDEAD ════════════════════
    {
        "tab": "Skeleton", "epithet": "Animated Bones",
        "role": "Minion — fragile but tireless",
        "size_type": "Medium undead, lawful evil",
        "ac": 13, "hp": 13, "speed": "30 ft.", "prof": 2, "init_mod": 2, "atk_per_round": 1,
        "scores": [10, 14, 15, 6, 8, 5], "mods": [0, 2, 2, -2, -1, -3],
        "save_prof": [False]*6, "saves": [0, 2, 2, -2, -1, -3],
        "resistances": "", "immunities": "poison", "cond_immune": "exhaustion, poisoned",
        "senses": "darkvision 60 ft., passive Perception 9", "vulnerable": "bludgeoning",
        "attacks": [
            ("Shortsword", "Melee Weapon", 4, "5 ft.", "1d6+2 piercing", "—", ""),
            ("Shortbow", "Ranged Weapon", 4, "80/320 ft.", "1d6+2 piercing", "—", ""),
        ],
        "traits": [
            ("Vulnerability", "Vulnerable to bludgeoning damage."),
        ],
        "bl_notes": "Undead — skeleton (CR 1/4)",
    },
    {
        "tab": "Zombie", "epithet": "Shambling Corpse",
        "role": "Minion — slow, refuses to die",
        "size_type": "Medium undead, neutral evil",
        "ac": 8, "hp": 22, "speed": "20 ft.", "prof": 2, "init_mod": -2, "atk_per_round": 1,
        "scores": [13, 6, 16, 3, 6, 5], "mods": [1, -2, 3, -4, -2, -3],
        "save_prof": [False, False, False, False, True, False], "saves": [1, -2, 3, -4, 0, -3],
        "resistances": "", "immunities": "poison", "cond_immune": "poisoned",
        "senses": "darkvision 60 ft., passive Perception 8",
        "attacks": [
            ("Slam", "Melee Weapon", 3, "5 ft.", "1d6+1 bludgeoning", "—", ""),
        ],
        "traits": [
            ("Undead Fortitude", "If reduced to 0 HP by damage that is not radiant or a critical hit, it makes a "
             "CON save (DC 5 + damage taken); on a success it drops to 1 HP instead."),
        ],
        "bl_notes": "Undead — zombie (CR 1/4)",
    },
    {
        "tab": "Shadow", "epithet": "Creeping Darkness",
        "role": "Ambusher — drains Strength, hides in dim light",
        "size_type": "Medium undead, chaotic evil",
        "ac": 12, "hp": 16, "speed": "40 ft.", "prof": 2, "init_mod": 2, "atk_per_round": 1,
        "scores": [6, 14, 13, 6, 10, 8], "mods": [-2, 2, 1, -2, 0, -1],
        "save_prof": [False]*6, "saves": [-2, 2, 1, -2, 0, -1],
        "resistances": "acid, cold, fire, lightning, thunder; " + NONMAG_BPS,
        "immunities": "necrotic, poison",
        "cond_immune": "exhaustion, frightened, grappled, paralyzed, petrified, poisoned, prone, restrained",
        "senses": "darkvision 60 ft., passive Perception 10", "vulnerable": "radiant",
        "attacks": [
            ("Strength Drain", "Melee Weapon", 4, "5 ft.", "2d6+2 necrotic", "—",
             "Target's STR reduced by 1d4 (restored on short/long rest). Dies if STR reaches 0; a humanoid that "
             "dies this way rises as a shadow 1d4 hours later."),
        ],
        "traits": [
            ("Amorphous", "Can move through a space as narrow as 1 inch without squeezing."),
            ("Shadow Stealth", "While in dim light or darkness, can take the Hide action as a bonus action."),
            ("Sunlight Weakness", "While in sunlight, has disadvantage on attack rolls, ability checks, and saving throws."),
            ("Vulnerability", "Vulnerable to radiant damage."),
        ],
        "bl_notes": "Undead — shadow (CR 1/2)",
    },
    {
        "tab": "Ghoul", "epithet": "Ravenous Dead",
        "role": "Striker — paralyzing claws",
        "size_type": "Medium undead, chaotic evil",
        "ac": 12, "hp": 22, "speed": "30 ft.", "prof": 2, "init_mod": 2, "atk_per_round": 1,
        "scores": [13, 15, 10, 7, 10, 6], "mods": [1, 2, 0, -2, 0, -2],
        "save_prof": [False]*6, "saves": [1, 2, 0, -2, 0, -2],
        "resistances": "", "immunities": "poison", "cond_immune": "charmed, exhaustion, poisoned",
        "senses": "darkvision 60 ft., passive Perception 10",
        "attacks": [
            ("Bite", "Melee Weapon", 2, "5 ft.", "2d6+2 piercing", "—", ""),
            ("Claws", "Melee Weapon", 4, "5 ft.", "2d4+2 slashing", "—",
             "Target (except elves and undead) makes a DC 10 CON save or is paralyzed for 1 min; repeats save at end of each of its turns."),
        ],
        "traits": [],
        "bl_notes": "Undead — ghoul (CR 1)",
    },
    {
        "tab": "Wight", "epithet": "Grave Lord",
        "role": "Elite — life-draining commander of the dead",
        "size_type": "Medium undead, neutral evil",
        "ac": 14, "hp": 45, "speed": "30 ft.", "prof": 2, "init_mod": 2, "atk_per_round": 2,
        "scores": [15, 14, 16, 10, 13, 15], "mods": [2, 2, 3, 0, 1, 2],
        "save_prof": [False]*6, "saves": [2, 2, 3, 0, 1, 2],
        "resistances": "necrotic; bludgeoning, piercing, slashing from nonmagical attacks that aren't silvered",
        "immunities": "poison", "cond_immune": "exhaustion, poisoned",
        "senses": "darkvision 60 ft., passive Perception 13",
        "attacks": [
            ("Longsword", "Melee Weapon", 4, "5 ft.", "1d8+2 slashing", "—", "1d10+2 if used two-handed."),
            ("Longbow", "Ranged Weapon", 4, "150/600 ft.", "1d8+2 piercing", "—", ""),
            ("Life Drain", "Melee Weapon", 4, "5 ft.", "1d6+2 necrotic", "DC 13 CON",
             "On a failed save, target's max HP is reduced by the damage taken (until a long rest). "
             "A humanoid slain this way rises as a zombie under the wight's control."),
        ],
        "traits": [
            ("Multiattack", "The wight makes two longsword/longbow attacks, or one Life Drain."),
            ("Sunlight Sensitivity", "While in sunlight, has disadvantage on attack rolls and Perception checks relying on sight."),
        ],
        "bl_notes": "Undead — wight (CR 3)",
    },
    {
        "tab": "Specter", "epithet": "Vengeful Spirit",
        "role": "Flying ambusher — passes through walls",
        "size_type": "Medium undead, chaotic evil",
        "ac": 12, "hp": 22, "speed": "0 ft., fly 50 ft. (hover)", "prof": 2, "init_mod": 2, "atk_per_round": 1,
        "scores": [1, 14, 11, 10, 10, 11], "mods": [-5, 2, 0, 0, 0, 0],
        "save_prof": [False]*6, "saves": [-5, 2, 0, 0, 0, 0],
        "resistances": "acid, cold, fire, lightning, thunder; bludgeoning, piercing, slashing from nonmagical attacks that aren't silvered",
        "immunities": "necrotic, poison",
        "cond_immune": "charmed, exhaustion, grappled, paralyzed, petrified, poisoned, prone, restrained, unconscious",
        "senses": "darkvision 60 ft., passive Perception 10",
        "attacks": [
            ("Life Drain", "Melee Spell", 4, "5 ft.", "3d6 necrotic", "DC 10 CON",
             "On a failed save, target's max HP is reduced by the damage taken (until a long rest)."),
        ],
        "traits": [
            ("Incorporeal Movement", "Can move through other creatures and objects as difficult terrain. Takes 1d10 force if it ends its turn inside an object."),
            ("Sunlight Sensitivity", "While in sunlight, has disadvantage on attack rolls and sight-based Perception checks."),
        ],
        "bl_notes": "Undead — specter (CR 1)",
    },
    {
        "tab": "Wraith", "epithet": "Shroud of Death",
        "role": "Elite — high-damage life drain, raises specters",
        "size_type": "Medium undead, neutral evil",
        "ac": 13, "hp": 67, "speed": "0 ft., fly 60 ft. (hover)", "prof": 3, "init_mod": 3, "atk_per_round": 1,
        "scores": [6, 16, 16, 12, 14, 15], "mods": [-2, 3, 3, 1, 2, 2],
        "save_prof": [False]*6, "saves": [-2, 3, 3, 1, 2, 2],
        "resistances": "acid, cold, fire, lightning, thunder; bludgeoning, piercing, slashing from nonmagical attacks that aren't silvered",
        "immunities": "necrotic, poison",
        "cond_immune": "charmed, exhaustion, grappled, paralyzed, petrified, poisoned, prone, restrained",
        "senses": "darkvision 60 ft., passive Perception 12",
        "attacks": [
            ("Life Drain", "Melee Spell", 6, "5 ft.", "4d8+3 necrotic", "DC 14 CON",
             "On a failed save, target's max HP is reduced by the damage taken (until a long rest). Dies if reduced to 0."),
        ],
        "traits": [
            ("Incorporeal Movement", "Can move through creatures and objects as difficult terrain; takes 1d10 force if it ends its turn inside an object."),
            ("Sunlight Sensitivity", "While in sunlight, has disadvantage on attack rolls and sight-based Perception checks."),
        ],
        "actions": [
            ("Create Specter", "Targets a humanoid that died within the last minute within 10 ft. The corpse rises as a "
             "specter under the wraith's control (up to seven at a time)."),
        ],
        "bl_notes": "Undead — wraith (CR 5)",
    },

    # ════════════════════ HUMANOID RAIDERS ════════════════════
    {
        "tab": "Goblin", "epithet": "Sneaky Raider",
        "role": "Minion skirmisher — darts in and out",
        "size_type": "Small humanoid (goblinoid), neutral evil",
        "ac": 15, "hp": 7, "speed": "30 ft.", "prof": 2, "init_mod": 2, "atk_per_round": 1,
        "scores": [8, 14, 10, 10, 8, 8], "mods": [-1, 2, 0, 0, -1, -1],
        "save_prof": [False]*6, "saves": [-1, 2, 0, 0, -1, -1],
        "resistances": "", "immunities": "", "cond_immune": "",
        "senses": "darkvision 60 ft., passive Perception 9",
        "attacks": [
            ("Scimitar", "Melee Weapon", 4, "5 ft.", "1d6+2 slashing", "—", ""),
            ("Shortbow", "Ranged Weapon", 4, "80/320 ft.", "1d6+2 piercing", "—", ""),
        ],
        "traits": [
            ("Nimble Escape", "Can take the Disengage or Hide action as a bonus action on each of its turns."),
        ],
        "bl_notes": "Humanoid — goblin (CR 1/4)",
    },
    {
        "tab": "Orc", "epithet": "Savage Warrior",
        "role": "Aggressive frontline — charges to close",
        "size_type": "Medium humanoid (orc), chaotic evil",
        "ac": 13, "hp": 15, "speed": "30 ft.", "prof": 2, "init_mod": 1, "atk_per_round": 1,
        "scores": [16, 12, 16, 7, 11, 10], "mods": [3, 1, 3, -2, 0, 0],
        "save_prof": [False]*6, "saves": [3, 1, 3, -2, 0, 0],
        "resistances": "", "immunities": "", "cond_immune": "",
        "senses": "darkvision 60 ft., passive Perception 10",
        "attacks": [
            ("Greataxe", "Melee Weapon", 5, "5 ft.", "1d12+3 slashing", "—", ""),
            ("Javelin", "Ranged Weapon", 5, "30/120 ft.", "1d6+3 piercing", "—", ""),
        ],
        "traits": [
            ("Aggressive", "As a bonus action, can move up to its speed toward a hostile creature it can see."),
        ],
        "bl_notes": "Humanoid — orc (CR 1/2)",
    },
    {
        "tab": "Ogre", "epithet": "Brute",
        "role": "Heavy hitter — big club, low cunning",
        "size_type": "Large giant, chaotic evil",
        "ac": 11, "hp": 59, "speed": "40 ft.", "prof": 2, "init_mod": -1, "atk_per_round": 1,
        "scores": [19, 8, 16, 5, 7, 7], "mods": [4, -1, 3, -3, -2, -2],
        "save_prof": [False]*6, "saves": [4, -1, 3, -3, -2, -2],
        "resistances": "", "immunities": "", "cond_immune": "",
        "senses": "darkvision 60 ft., passive Perception 8",
        "attacks": [
            ("Greatclub", "Melee Weapon", 6, "5 ft.", "2d8+4 bludgeoning", "—", ""),
            ("Javelin", "Ranged Weapon", 6, "30/120 ft.", "2d6+4 piercing", "—", ""),
        ],
        "traits": [],
        "bl_notes": "Giant — ogre (CR 2)",
    },
    {
        "tab": "Bugbear", "epithet": "Goblinoid Bully",
        "role": "Ambusher — devastating surprise attacks",
        "size_type": "Medium humanoid (goblinoid), chaotic evil",
        "ac": 16, "hp": 27, "speed": "30 ft.", "prof": 2, "init_mod": 2, "atk_per_round": 1,
        "scores": [15, 14, 13, 8, 11, 9], "mods": [2, 2, 1, -1, 0, -1],
        "save_prof": [False]*6, "saves": [2, 2, 1, -1, 0, -1],
        "resistances": "", "immunities": "", "cond_immune": "",
        "senses": "darkvision 60 ft., passive Perception 10",
        "attacks": [
            ("Morningstar", "Melee Weapon", 4, "5 ft.", "2d8+2 piercing", "—", "Includes +1d8 Brute die."),
            ("Javelin", "Ranged Weapon", 4, "30/120 ft.", "1d6+2 piercing", "—", "2d6+2 in melee (Brute)."),
        ],
        "traits": [
            ("Brute", "A melee weapon deals one extra die of its damage (already included)."),
            ("Surprise Attack", "If it surprises a creature and hits it in the first round, the attack deals an extra 2d6 damage."),
        ],
        "bl_notes": "Humanoid — bugbear (CR 1)",
    },
    {
        "tab": "Orc War Chief", "epithet": "Warband Leader",
        "role": "Elite commander — rallies the warband",
        "size_type": "Medium humanoid (orc), chaotic evil",
        "ac": 16, "hp": 93, "speed": "30 ft.", "prof": 2, "init_mod": 1, "atk_per_round": 2,
        "scores": [18, 12, 18, 11, 11, 16], "mods": [4, 1, 4, 0, 0, 3],
        "save_prof": [True, False, True, False, True, False], "saves": [6, 1, 6, 0, 2, 3],
        "resistances": "", "immunities": "", "cond_immune": "",
        "senses": "darkvision 60 ft., passive Perception 10",
        "attacks": [
            ("Greataxe", "Melee Weapon", 6, "5 ft.", "1d12+4 slashing plus 1d8 necrotic", "—",
             "Extra 1d8 from Gruumsh's Fury (already included)."),
            ("Spear", "Melee Weapon", 6, "5 ft.", "1d6+4 piercing plus 1d8 necrotic", "—", "Thrown 20/60 ft."),
        ],
        "traits": [
            ("Multiattack", "The war chief makes two greataxe attacks."),
            ("Aggressive", "As a bonus action, can move up to its speed toward a hostile creature it can see."),
            ("Gruumsh's Fury", "Deals an extra 1d8 damage with weapon attacks (already included)."),
        ],
        "actions": [
            ("Battle Cry (1/Day)", "As a bonus action, each creature of the war chief's choice within 30 ft. that can "
             "hear it gains advantage on attack rolls until the start of the war chief's next turn. The war chief can then make one attack as a bonus action."),
        ],
        "bl_notes": "Humanoid — orc war chief (CR 4) — PRIORITY TARGET",
    },

    # ════════════════════ GIANTS ════════════════════
    {
        "tab": "Stone Giant", "epithet": "Mountain Sentinel",
        "role": "Artillery — hurls boulders, knocks prone",
        "size_type": "Huge giant, neutral",
        "ac": 17, "hp": 126, "speed": "40 ft.", "prof": 3, "init_mod": 2, "atk_per_round": 2,
        "scores": [23, 15, 20, 10, 12, 9], "mods": [6, 2, 5, 0, 1, -1],
        "save_prof": [False, True, True, False, True, False], "saves": [6, 5, 8, 0, 4, -1],
        "resistances": "", "immunities": "", "cond_immune": "",
        "senses": "darkvision 60 ft., passive Perception 14",
        "attacks": [
            ("Greatclub", "Melee Weapon", 9, "15 ft.", "3d8+6 bludgeoning", "—", ""),
            ("Rock", "Ranged Weapon", 9, "60/240 ft.", "4d10+6 bludgeoning", "—",
             "If a Large or smaller creature is hit, DC 17 STR save or knocked prone."),
        ],
        "traits": [
            ("Multiattack", "The giant makes two greatclub attacks."),
            ("Rock Catching", "If a rock or similar object is hurled at it, it can catch it with a DC 10+ reaction."),
            ("Stone Camouflage", "Advantage on Stealth checks made to hide in rocky terrain."),
        ],
        "bl_notes": "Giant — stone (CR 7)",
    },
    {
        "tab": "Frost Giant", "epithet": "Ice Reaver",
        "role": "Heavy melee — great axe and boulders",
        "size_type": "Huge giant, neutral evil",
        "ac": 15, "hp": 138, "speed": "40 ft.", "prof": 3, "init_mod": -1, "atk_per_round": 2,
        "scores": [23, 9, 21, 9, 10, 12], "mods": [6, -1, 5, -1, 0, 1],
        "save_prof": [False, False, True, False, True, True], "saves": [6, -1, 8, -1, 3, 4],
        "resistances": "", "immunities": "cold", "cond_immune": "",
        "senses": "passive Perception 13",
        "attacks": [
            ("Greataxe", "Melee Weapon", 9, "10 ft.", "3d12+6 slashing", "—", ""),
            ("Rock", "Ranged Weapon", 9, "60/240 ft.", "4d10+6 bludgeoning", "—", ""),
        ],
        "traits": [
            ("Multiattack", "The giant makes two greataxe attacks."),
        ],
        "bl_notes": "Giant — frost/ice (CR 8)",
    },
    {
        "tab": "Fire Giant", "epithet": "Forge Lord",
        "role": "Disciplined heavy — greatsword juggernaut",
        "size_type": "Huge giant, lawful evil",
        "ac": 18, "hp": 162, "speed": "30 ft.", "prof": 4, "init_mod": -1, "atk_per_round": 2,
        "scores": [25, 9, 23, 10, 14, 13], "mods": [7, -1, 6, 0, 2, 1],
        "save_prof": [False, True, True, False, True, False], "saves": [7, 3, 10, 0, 6, 1],
        "resistances": "", "immunities": "fire", "cond_immune": "",
        "senses": "passive Perception 16",
        "attacks": [
            ("Greatsword", "Melee Weapon", 11, "10 ft.", "6d6+7 slashing", "—", ""),
            ("Rock", "Ranged Weapon", 11, "60/240 ft.", "4d10+7 bludgeoning", "—", ""),
        ],
        "traits": [
            ("Multiattack", "The giant makes two greatsword attacks."),
        ],
        "bl_notes": "Giant — fire (CR 9)",
    },
    {
        "tab": "Storm Giant", "epithet": "Tempest Sovereign",
        "role": "Apex caster-warrior — lightning from afar",
        "size_type": "Huge giant, chaotic good",
        "ac": 16, "hp": 230, "speed": "50 ft., swim 50 ft.", "prof": 5, "init_mod": 2, "atk_per_round": 2,
        "scores": [29, 14, 20, 16, 18, 18], "mods": [9, 2, 5, 3, 4, 4],
        "save_prof": [True, False, True, False, True, True], "saves": [14, 2, 10, 3, 9, 9],
        "resistances": "cold", "immunities": "lightning, thunder", "cond_immune": "",
        "senses": "darkvision 120 ft., passive Perception 17",
        "attacks": [
            ("Greatsword", "Melee Weapon", 14, "10 ft.", "6d6+9 slashing", "—", ""),
            ("Rock", "Ranged Weapon", 14, "60/240 ft.", "4d12+9 bludgeoning", "—", ""),
            ("Lightning Strike", "Special (Recharge 5-6)", "—", "500 ft.", "12d8 lightning", "DC 17 DEX",
             "Hurls a lightning bolt at a point; each creature within 10 ft. takes damage (half on save)."),
        ],
        "traits": [
            ("Multiattack", "The giant makes two greatsword attacks."),
            ("Amphibious", "Can breathe air and water."),
            ("Innate Spellcasting", "Can innately cast Control Weather, Detect Magic, Feather Fall, Levitate, Light, and (1/day each) Control Water and Water Breathing."),
        ],
        "bl_notes": "Giant — storm (CR 13)",
    },

    # ════════════════════ ADULT CHROMATIC DRAGONS ════════════════════
    {
        "tab": "Adult Red Dragon", "epithet": "Tyrant of Flame",
        "role": "Apex solo — fire breath, frightful presence, legendary",
        "size_type": "Huge dragon, chaotic evil",
        "ac": 19, "hp": 256, "speed": "40 ft., climb 40 ft., fly 80 ft.", "prof": 6, "init_mod": 0, "atk_per_round": 3,
        "scores": [27, 10, 25, 16, 13, 21], "mods": [8, 0, 7, 3, 1, 5],
        "save_prof": [False, True, True, False, True, True], "saves": [8, 6, 13, 3, 7, 11],
        "resistances": "", "immunities": "fire", "cond_immune": "",
        "senses": "blindsight 60 ft., darkvision 120 ft., passive Perception 23",
        "attacks": [
            ("Bite", "Melee Weapon", 14, "10 ft.", "2d10+8 piercing plus 2d6 fire", "—", ""),
            ("Claw", "Melee Weapon", 14, "5 ft.", "2d6+8 slashing", "—", ""),
            ("Tail", "Melee Weapon", 14, "15 ft.", "2d8+8 bludgeoning", "—", ""),
            ("Fire Breath", "Special (Recharge 5-6)", "—", "60-ft cone", "16d6 fire", "DC 21 DEX", "Half on save."),
            ("Frightful Presence", "Special", "—", "120 ft.", "—", "DC 19 WIS",
             "Each chosen creature that fails is frightened for 1 min (repeat save at end of each of its turns; success ends and grants immunity for 24 hrs)."),
        ],
        "traits": [
            ("Multiattack", "Frightful Presence, then one Bite and two Claws."),
            ("Legendary Resistance (3/Day)", "If it fails a saving throw, it can choose to succeed instead."),
            ("Legendary Actions (3/round)", "Detect (Perception check); Tail Attack; Wing Attack (costs 2): "
             "each creature within 10 ft. makes a DC 22 DEX save, taking 2d6+8 bludgeoning and knocked prone on a fail; the dragon then flies up to half its speed."),
        ],
        "bl_notes": "Dragon — adult red (CR 17) — BOSS",
    },
    {
        "tab": "Adult Blue Dragon", "epithet": "Storm of the Desert",
        "role": "Apex solo — lightning line, burrows, legendary",
        "size_type": "Huge dragon, lawful evil",
        "ac": 19, "hp": 225, "speed": "40 ft., burrow 30 ft., fly 80 ft.", "prof": 6, "init_mod": 0, "atk_per_round": 3,
        "scores": [25, 10, 23, 16, 15, 19], "mods": [7, 0, 6, 3, 2, 4],
        "save_prof": [False, True, True, False, True, True], "saves": [7, 6, 12, 3, 8, 10],
        "resistances": "", "immunities": "lightning", "cond_immune": "",
        "senses": "blindsight 60 ft., darkvision 120 ft., passive Perception 22",
        "attacks": [
            ("Bite", "Melee Weapon", 12, "10 ft.", "2d10+7 piercing plus 1d10 lightning", "—", ""),
            ("Claw", "Melee Weapon", 12, "5 ft.", "2d6+7 slashing", "—", ""),
            ("Tail", "Melee Weapon", 12, "15 ft.", "2d8+7 bludgeoning", "—", ""),
            ("Lightning Breath", "Special (Recharge 5-6)", "—", "90-ft line (5 ft wide)", "12d10 lightning", "DC 19 DEX", "Half on save."),
            ("Frightful Presence", "Special", "—", "120 ft.", "—", "DC 17 WIS",
             "Each chosen creature that fails is frightened for 1 min (repeat save each turn)."),
        ],
        "traits": [
            ("Multiattack", "Frightful Presence, then one Bite and two Claws."),
            ("Legendary Resistance (3/Day)", "If it fails a saving throw, it can choose to succeed instead."),
            ("Legendary Actions (3/round)", "Detect; Tail Attack; Wing Attack (costs 2): DC 20 DEX, 2d6+7 bludgeoning and prone, then fly half speed."),
        ],
        "bl_notes": "Dragon — adult blue (CR 16) — BOSS",
    },
    {
        "tab": "Adult Green Dragon", "epithet": "Manipulator of the Wood",
        "role": "Apex solo — poison breath, cunning, legendary",
        "size_type": "Huge dragon, lawful evil",
        "ac": 19, "hp": 207, "speed": "40 ft., fly 80 ft., swim 40 ft.", "prof": 5, "init_mod": 1, "atk_per_round": 3,
        "scores": [23, 12, 21, 18, 15, 17], "mods": [6, 1, 5, 4, 2, 3],
        "save_prof": [False, True, True, False, True, True], "saves": [6, 6, 10, 4, 7, 8],
        "resistances": "", "immunities": "poison", "cond_immune": "poisoned",
        "senses": "blindsight 60 ft., darkvision 120 ft., passive Perception 22",
        "attacks": [
            ("Bite", "Melee Weapon", 11, "10 ft.", "2d10+6 piercing plus 2d6 poison", "—", ""),
            ("Claw", "Melee Weapon", 11, "5 ft.", "2d6+6 slashing", "—", ""),
            ("Tail", "Melee Weapon", 11, "15 ft.", "2d8+6 bludgeoning", "—", ""),
            ("Poison Breath", "Special (Recharge 5-6)", "—", "60-ft cone", "16d6 poison", "DC 18 CON", "Half on save."),
            ("Frightful Presence", "Special", "—", "120 ft.", "—", "DC 16 WIS",
             "Each chosen creature that fails is frightened for 1 min (repeat save each turn)."),
        ],
        "traits": [
            ("Multiattack", "Frightful Presence, then one Bite and two Claws."),
            ("Amphibious", "Can breathe air and water."),
            ("Legendary Resistance (3/Day)", "If it fails a saving throw, it can choose to succeed instead."),
            ("Legendary Actions (3/round)", "Detect; Tail Attack; Wing Attack (costs 2): DC 19 DEX, 2d6+6 bludgeoning and prone, then fly half speed."),
        ],
        "bl_notes": "Dragon — adult green (CR 15) — BOSS",
    },
    {
        "tab": "Adult Black Dragon", "epithet": "Lord of the Swamp",
        "role": "Apex solo — acid line, lurks in water, legendary",
        "size_type": "Huge dragon, chaotic evil",
        "ac": 19, "hp": 195, "speed": "40 ft., fly 80 ft., swim 40 ft.", "prof": 5, "init_mod": 2, "atk_per_round": 3,
        "scores": [23, 14, 21, 14, 13, 17], "mods": [6, 2, 5, 2, 1, 3],
        "save_prof": [False, True, True, False, True, True], "saves": [6, 7, 10, 2, 6, 8],
        "resistances": "", "immunities": "acid", "cond_immune": "",
        "senses": "blindsight 60 ft., darkvision 120 ft., passive Perception 21",
        "attacks": [
            ("Bite", "Melee Weapon", 11, "10 ft.", "2d10+6 piercing plus 1d8 acid", "—", ""),
            ("Claw", "Melee Weapon", 11, "5 ft.", "2d6+6 slashing", "—", ""),
            ("Tail", "Melee Weapon", 11, "15 ft.", "2d8+6 bludgeoning", "—", ""),
            ("Acid Breath", "Special (Recharge 5-6)", "—", "60-ft line (5 ft wide)", "12d8 acid", "DC 18 DEX", "Half on save."),
            ("Frightful Presence", "Special", "—", "120 ft.", "—", "DC 16 WIS",
             "Each chosen creature that fails is frightened for 1 min (repeat save each turn)."),
        ],
        "traits": [
            ("Multiattack", "Frightful Presence, then one Bite and two Claws."),
            ("Amphibious", "Can breathe air and water."),
            ("Legendary Resistance (3/Day)", "If it fails a saving throw, it can choose to succeed instead."),
            ("Legendary Actions (3/round)", "Detect; Tail Attack; Wing Attack (costs 2): DC 19 DEX, 2d6+6 bludgeoning and prone, then fly half speed."),
        ],
        "bl_notes": "Dragon — adult black (CR 14) — BOSS",
    },
    {
        "tab": "Adult White Dragon", "epithet": "Fury of the Frozen North",
        "role": "Apex solo — cold breath, savage, legendary",
        "size_type": "Huge dragon, chaotic evil",
        "ac": 18, "hp": 200, "speed": "40 ft., burrow 30 ft., fly 80 ft., swim 40 ft.", "prof": 5, "init_mod": 0, "atk_per_round": 3,
        "scores": [22, 10, 22, 8, 12, 12], "mods": [6, 0, 6, -1, 1, 1],
        "save_prof": [False, True, True, False, True, True], "saves": [6, 5, 11, -1, 6, 6],
        "resistances": "", "immunities": "cold", "cond_immune": "",
        "senses": "blindsight 60 ft., darkvision 120 ft., passive Perception 21",
        "attacks": [
            ("Bite", "Melee Weapon", 11, "10 ft.", "2d10+6 piercing plus 1d8 cold", "—", ""),
            ("Claw", "Melee Weapon", 11, "5 ft.", "2d6+6 slashing", "—", ""),
            ("Tail", "Melee Weapon", 11, "15 ft.", "2d8+6 bludgeoning", "—", ""),
            ("Cold Breath", "Special (Recharge 5-6)", "—", "60-ft cone", "12d8 cold", "DC 19 CON", "Half on save."),
            ("Frightful Presence", "Special", "—", "120 ft.", "—", "DC 14 WIS",
             "Each chosen creature that fails is frightened for 1 min (repeat save each turn)."),
        ],
        "traits": [
            ("Multiattack", "Frightful Presence, then one Bite and two Claws."),
            ("Ice Walk", "Can move across icy/snowy difficult terrain without extra cost."),
            ("Legendary Resistance (3/Day)", "If it fails a saving throw, it can choose to succeed instead."),
            ("Legendary Actions (3/round)", "Detect; Tail Attack; Wing Attack (costs 2): DC 19 DEX, 2d6+6 bludgeoning and prone, then fly half speed."),
        ],
        "bl_notes": "Dragon — adult white (CR 13) — BOSS",
    },
]

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if not os.path.exists(WORKBOOK):
        print(f"ERROR: Workbook not found: {WORKBOOK}")
        sys.exit(1)

    print(f"Loading {WORKBOOK} ...")
    shutil.copy2(WORKBOOK, TMP)
    wb = openpyxl.load_workbook(TMP)

    existing = set(wb.sheetnames)
    print(f"Existing sheets ({len(existing)}): {sorted(existing)}")

    for m in MONSTERS:
        add_creature_tab(wb, m)
        row = add_battle_list_row(wb, m["tab"], qty=0, notes=m.get("bl_notes", ""))
        print(f"  + tab '{m['tab']}'  (Battle List row {row})")

    wb.save(TMP)
    shutil.copy2(TMP, WORKBOOK)

    final = openpyxl.load_workbook(WORKBOOK).sheetnames
    print(f"\nDone. {len(MONSTERS)} monster tabs added.")
    print(f"Workbook now has {len(final)} sheets.")
    # Safety check: every previously-existing sheet must still be present.
    missing = existing - set(final)
    print("Lost sheets:" , missing if missing else "NONE (all preserved)")
